"""
Extended sink/output nodes for F-Pulse.

Covers: CSV, JSON, Excel, S3/MinIO, Kafka, REST API, Webhook, Email, Delta, Warehouse.
All sink nodes return the input data as-is for downstream preview.
"""

from __future__ import annotations

import json
import os
import smtplib
import tempfile
import urllib.request
import urllib.error
from email.message import EmailMessage
from typing import Any, TYPE_CHECKING

# Stage 2.5b: duckdb only used for type annotations (execute() returns,
# helper signatures). Runtime data flow is through ctx.conn.
if TYPE_CHECKING:
    import duckdb

from fpulse.ir.schema import StepType
from fpulse.intelligence.schema_policy import (
    DEFAULT_POLICY,
    SchemaPolicy,
    evaluate_policy,
    schema_policy_param,
)
from fpulse.nodes.base import BaseNode, ExecutionContext
from fpulse.nodes.registry import register
from fpulse.sinks.dedupe_store import (
    DEFAULT_TTL_SECONDS as _IDEMP_DEFAULT_TTL,
    get_dedupe_store as _get_dedupe_store,
)
from fpulse.sinks.idempotency_helper import should_skip as _should_skip
from fpulse.security.ssrf import (
    check_url as _ssrf_check_url,
    SsrfBlockedError as _SsrfBlockedError,
)


# SSRF + redirect hardening for sink HTTP calls. An API/webhook sink POSTs the
# dataset OUT, so an unvalidated or redirected target is a data-exfiltration
# vector. Validate the URL up front, then refuse to auto-follow 3xx — plain
# urlopen follows a redirect to an unvalidated Location and re-resolves DNS,
# defeating the up-front check (redirect-to-internal / DNS-rebinding bypass).
# Env escape hatch mirrors the API Source / HTTP Request nodes.
_HTTP_ALLOW_PRIVATE_ENV = "FPULSE_HTTP_ALLOW_PRIVATE"


class _NoRedirectHandler(urllib.request.HTTPRedirectHandler):
    def redirect_request(self, req, fp, code, msg, headers, newurl):  # noqa: D401
        raise _SsrfBlockedError(
            f"Refusing to follow redirect ({code}) to {newurl!r} — "
            f"redirects are blocked to prevent SSRF/DNS-rebinding bypass."
        )


def _no_redirect_opener() -> "urllib.request.OpenerDirector":
    return urllib.request.build_opener(_NoRedirectHandler())


# ── Idempotency-key params shared by every external sink ─────────────
#
# Two parameters that every external sink (email, webhook, api, kafka,
# slack) exposes via param_schema(). Centralised so the UI labels and
# descriptions stay identical across sinks — users learn the concept
# once and recognise the same control on every sink card.
_IDEMPOTENCY_KEY_PARAM = {
    "name": "idempotency_key",
    "type": "text",
    "label": "Idempotency Key Template",
    "tier": "optional",
    "placeholder": "{user_id}|{event_id}",
    "description": (
        "Per-row template that uniquely identifies a side-effect target. "
        "Supports {column_name} substitution. When set, rows whose hash "
        "has already been sent for this sink (within the TTL) are "
        "SKIPPED on re-run, retry, or backfill. Leave empty to keep "
        "today's no-dedupe behaviour."
    ),
}

_IDEMPOTENCY_TTL_PARAM = {
    "name": "idempotency_ttl_days",
    "type": "number",
    "label": "Idempotency TTL (days)",
    "tier": "optional",
    "default": 30,
    "description": (
        "How long the dedupe marker is honoured before the row can be "
        "re-sent. Only used when Idempotency Key Template is set."
    ),
}


def _partition_for_idempotency(
    ctx: ExecutionContext,
    params: dict,
    records: list[dict],
) -> tuple[list[dict], int, str, str]:
    """Split ``records`` into (to_send, skipped_count, pipeline_id, sink_step_id).

    Centralised pre-fire filter used by every external sink. When the
    user has set ``idempotency_key`` on the sink, we:

      1. Resolve the dedupe store from the module singleton (wired to
         the SQLite DB at app startup; defensive no-op in tests that
         don't wire it).
      2. Render + sha256 the key template against each row.
      3. Ask the store ``seen()`` for each hash; rows that come back
         True are dropped from the to_send list.

    Returns the rows we should actually fire side effects for, plus a
    count of how many were skipped — the caller folds the count into
    its return log so the run history shows "Sent 8 / Skipped 12".

    When ``idempotency_key`` is empty/unset, returns ``(records, 0,
    pipeline_id, sink_step_id)`` — i.e. no rows are dropped, and the
    caller sees behaviour identical to today.

    The pipeline_id + sink_step_id are also returned so the caller can
    pass them straight to ``dedupe_store.record(...)`` for each row it
    successfully sends, without re-deriving them from ctx + params.
    """
    key_expression = (params.get("idempotency_key") or "").strip()
    pipeline_id = ctx.workflow_id or ""
    sink_step_id = params.get("_step_id", "") or ""

    if not key_expression:
        return list(records), 0, pipeline_id, sink_step_id

    store = _get_dedupe_store()
    to_send: list[dict] = []
    skipped = 0
    for row in records:
        skip, _hash = _should_skip(
            pipeline_id=pipeline_id,
            sink_step_id=sink_step_id,
            row=row,
            key_expression=key_expression,
            dedupe_store=store,
        )
        if skip:
            skipped += 1
        else:
            to_send.append(row)
    return to_send, skipped, pipeline_id, sink_step_id


def _record_idempotency(
    params: dict,
    pipeline_id: str,
    sink_step_id: str,
    row: dict,
) -> None:
    """Best-effort: write the dedupe marker for a row we just sent.

    Called once per successfully-fired side effect. Hashes the row +
    template the same way ``_partition_for_idempotency`` did the
    seen() check so the stored key matches the lookup key exactly.

    No-op when ``idempotency_key`` is empty (the per-row dedup wasn't
    in play in the first place). All failures are swallowed by the
    underlying store — a marker write that fails simply means the next
    run may send a duplicate, which is the same failure mode as
    before this feature existed.
    """
    key_expression = (params.get("idempotency_key") or "").strip()
    if not key_expression:
        return
    from fpulse.sinks.idempotency_helper import compute_row_hash
    key_hash = compute_row_hash(row, key_expression)
    if not key_hash:
        return
    ttl_days = params.get("idempotency_ttl_days", 30)
    try:
        ttl_seconds = int(ttl_days) * 86400 if ttl_days else _IDEMP_DEFAULT_TTL
    except (TypeError, ValueError):
        ttl_seconds = _IDEMP_DEFAULT_TTL
    _get_dedupe_store().record(
        pipeline_id=pipeline_id,
        sink_step_id=sink_step_id,
        key_hash=key_hash,
        ttl_seconds=ttl_seconds,
    )


# ── Helpers ──

def _get_input(ctx: ExecutionContext, params: dict) -> duckdb.DuckDBPyRelation:
    """Get the first input relation or raise."""
    inputs = ctx.get_inputs(params.get("_input_step_ids", []))
    if not inputs:
        raise ValueError("Sink node has no input data")
    return inputs[0]


def _resolve_output_path(file_path: str, data_dir: str, default_name: str) -> str:
    """Resolve output file path, creating directories as needed."""
    if not file_path:
        file_path = os.path.join(data_dir, default_name)
    elif not os.path.isabs(file_path):
        file_path = os.path.join(data_dir, file_path)
    os.makedirs(os.path.dirname(file_path) or ".", exist_ok=True)
    return file_path


def _register_output_in_storage_index(
    ctx: ExecutionContext,
    absolute_path: str,
    file_format: str,
) -> None:
    """Insert a storage_objects row for a sink-produced file (P0 Day 4, 2026-05-23).

    Previously the index was back-filled by the reconciler on the NEXT
    boot, so the Storage page's Pipeline Outputs tab would show stale
    data immediately after a run. This helper inserts the row at write
    time so the file appears the moment the sink finishes.

    Idempotent: if a row already exists for the same path, the size +
    updated_at are refreshed instead of duplicating. Reconciler stays
    as fallback for files written by external processes or older
    builds that don't carry this hook.

    Best-effort — any failure (workspace_id unresolvable, datastore
    not initialised in this process, path outside data_dir) is swallowed
    so a registration glitch can never break a successful sink write.
    """
    try:
        import os as _os
        from datetime import datetime, timezone

        # data_dir-relative path. If the sink wrote outside data_dir
        # (custom file_path the user set), skip — only sandboxed outputs
        # belong in the workspace index.
        data_dir = ctx.data_dir
        abs_dir = _os.path.abspath(data_dir)
        abs_path = _os.path.abspath(absolute_path)
        if not abs_path.startswith(abs_dir + _os.sep):
            return
        rel_path = _os.path.relpath(abs_path, abs_dir).replace("\\", "/")

        # Workspace + workflow context come from ctx. The workflow_ref
        # is stamped by the executor before the run; workspace_id lives
        # on the workflow IR.
        workflow = getattr(ctx, "_workflow_ref", None)
        workspace_id = (
            (getattr(workflow, "workspace_id", None) if workflow else None)
            or "default"
        )
        pipeline_id = getattr(workflow, "id", None) if workflow else None
        run_id = getattr(ctx, "run_id", None)

        # Datastore is in app_state. If app_state isn't wired (some test
        # paths instantiate ExecutionContext directly), skip — reconciler
        # picks it up later.
        from fpulse.datastore.store import get_store as get_datastore
        from fpulse.datastore.models import StorageObject, OBJECT_KIND_OUTPUT

        store = get_datastore()
        if store is None:
            return

        try:
            size_bytes = _os.path.getsize(abs_path)
        except OSError:
            size_bytes = 0
        name = _os.path.basename(abs_path)

        # Idempotent upsert by path + workspace. If a row already exists
        # for this path, refresh size + updated_at (this run overwrote
        # the file).
        existing = None
        try:
            for obj in store.list_objects(workspace_id, include_deleted=False):
                if obj.path == rel_path:
                    existing = obj
                    break
        except Exception:
            existing = None

        now = datetime.now(timezone.utc)
        if existing:
            existing.size_bytes = size_bytes
            existing.updated_at = now
            if pipeline_id and not existing.pipeline_id:
                existing.pipeline_id = pipeline_id
            if run_id and not existing.run_id:
                existing.run_id = run_id
            store.save_object(existing)
        else:
            row = StorageObject(
                workspace_id=workspace_id,
                kind=OBJECT_KIND_OUTPUT,
                name=name,
                path=rel_path,
                format=(file_format or "").lower() or None,
                size_bytes=size_bytes,
                pipeline_id=pipeline_id,
                run_id=run_id,
                created_at=now,
                updated_at=now,
            )
            store.save_object(row)
    except Exception:
        # Never raise — sink write already succeeded; index registration
        # is a best-effort enrichment, reconciler is the fallback.
        import logging
        logging.getLogger(__name__).debug(
            "register_output_in_storage_index: best-effort path failed",
            exc_info=True,
        )


def _relation_to_dicts(conn: duckdb.DuckDBPyConnection,
                        relation: duckdb.DuckDBPyRelation,
                        table_alias: str = "__sink_export") -> list[dict]:
    """Convert a DuckDB relation to a list of dicts."""
    conn.register(table_alias, relation)
    columns = relation.columns
    rows = conn.sql(f"SELECT * FROM {table_alias}").fetchall()
    return [dict(zip(columns, row)) for row in rows]


# ── Schema-policy helpers (2026-05-27) ───────────────────────────────────
#
# WarehouseSink uses these to consult schema_policy.evaluate_policy()
# before issuing ALTER DDL. The helpers live here rather than in
# schema_policy.py because they're sink-side concerns (relation +
# information_schema introspection); the policy module stays pure.

def _resolve_schema_policy(params: dict) -> str:
    """Map ``schema_policy`` + legacy ``auto_evolve`` to a single value.

    Existing pipelines have ``auto_evolve=True`` (the previous default
    that wired the add-column path) or ``auto_evolve=False`` (which
    didn't quite mean "strict" — the column-add branch just didn't
    run, leaving INSERT to fail noisily on a column count mismatch).
    We treat that legacy combination as the new explicit policies:

        auto_evolve=True   → add_columns   (functionally identical)
        auto_evolve=False  → strict        (cleaner failure mode)

    A workflow that ALREADY sets ``schema_policy`` wins over the
    legacy flag — that's how the migration story works: edit the
    pipeline once, the legacy flag becomes a no-op.
    """
    explicit = params.get("schema_policy")
    if explicit:
        return str(explicit)
    legacy = params.get("auto_evolve")
    if legacy is None:
        return DEFAULT_POLICY.value
    return (
        SchemaPolicy.ADD_COLUMNS.value if legacy
        else SchemaPolicy.STRICT.value
    )


def _publish_warehouse_drift_event(
    ctx: ExecutionContext,
    *,
    step_id: str,
    table_name: str,
    decision,
    applied: bool,
    rejection_reason: str = "",
) -> None:
    """Best-effort publish of SchemaDriftDetected for a warehouse-sink write.

    Warehouse sinks don't write to a managed table, so the event
    carries an empty ``table_id`` and the display name only. The
    schema_history append is skipped (history is per-managed-table);
    the event is the only audit trail for warehouse drift.
    """
    try:
        from fpulse.events import SchemaDriftDetected, get_event_bus
        bus = get_event_bus()
        if bus is None:
            return
        summary = decision.to_summary()
        bus.publish(SchemaDriftDetected(
            run_id=getattr(ctx, "run_id", "") or "",
            step_id=step_id or "",
            workspace_id=getattr(ctx, "workspace_id", "default") or "default",
            table_id="",
            table_name=table_name,
            policy=decision.policy.value,
            severity=decision.severity if applied else "critical",
            applied=applied,
            schema_version=0,
            added_columns=summary.get("added") or [],
            dropped_columns=summary.get("dropped") or [],
            type_changes=summary.get("type_changed") or [],
            rejection_reason=rejection_reason,
        ))
    except Exception:
        import logging
        logging.getLogger(__name__).debug(
            "warehouse drift event publish skipped", exc_info=True,
        )


# ── CSV Sink ──

@register(StepType.CSV_SINK)
class CsvSinkNode(BaseNode):
    """Write data to a CSV file using DuckDB COPY."""
    display_name = "CSV Sink"
    category = "output"
    description = "Write data to a CSV file"

    def execute(self, ctx: ExecutionContext) -> duckdb.DuckDBPyRelation:
        source = _get_input(ctx, self.params)
        file_path = _resolve_output_path(
            self.params.get("file_path", ""), ctx.data_dir, "output.csv"
        )
        delimiter = self.params.get("delimiter", ",")
        header = self.params.get("header", True)
        quote_all = self.params.get("quote_all", False)

        ctx.conn.register("__csv_sink_data", source)

        header_opt = "HEADER" if header else "HEADER false"
        quote_opt = ", FORCE_QUOTE *" if quote_all else ""
        ctx.conn.sql(
            f"COPY __csv_sink_data TO '{file_path}' "
            f"(FORMAT CSV, DELIMITER '{delimiter}', {header_opt}{quote_opt})"
        )

        # P0 Day 4 (2026-05-23) — register the output in the storage
        # index immediately so the Storage → Pipeline Outputs tab shows
        # this row without waiting for the next-boot reconciler sweep.
        _register_output_in_storage_index(ctx, file_path, "csv")

        return source

    @staticmethod
    def default_params() -> dict[str, Any]:
        return {"file_path": "", "delimiter": ",", "header": True, "quote_all": False}

    @staticmethod
    def preview_message(params, row_count):
        # X4 (2026-05-30) — observability-only override of the generic
        # "side effect skipped" message. Tells the operator what file
        # the dry-run WOULD have written.
        target = params.get("file_path") or "(default output path)"
        return f"would write {row_count} row{'s' if row_count != 1 else ''} to {target}"

    @staticmethod
    def param_schema() -> list[dict]:
        return [
            {"name": "file_path", "type": "text", "label": "File Path",
             "placeholder": "output/results.csv",
             "description": "Leave empty for default output path."},
            {"name": "delimiter", "type": "select", "label": "Delimiter",
             "options": [",", ";", "\\t", "|"], "default": ","},
            {"name": "header", "type": "boolean", "label": "Include Header", "default": True},
            {"name": "quote_all", "type": "boolean", "label": "Quote All Fields", "default": False},
        ]


# ── JSON Sink ──

@register(StepType.JSON_SINK)
class JsonSinkNode(BaseNode):
    """Write data to a JSON file using DuckDB COPY."""
    display_name = "JSON Sink"
    category = "output"
    description = "Write data to a JSON file"

    def execute(self, ctx: ExecutionContext) -> duckdb.DuckDBPyRelation:
        source = _get_input(ctx, self.params)
        file_path = _resolve_output_path(
            self.params.get("file_path", ""), ctx.data_dir, "output.json"
        )
        json_format = self.params.get("format", "array")

        ctx.conn.register("__json_sink_data", source)

        if json_format == "lines":
            # Newline-delimited JSON (each row is a JSON object on its own line)
            ctx.conn.sql(
                f"COPY __json_sink_data TO '{file_path}' (FORMAT JSON, ARRAY false)"
            )
        else:
            # JSON array (default)
            ctx.conn.sql(
                f"COPY __json_sink_data TO '{file_path}' (FORMAT JSON, ARRAY true)"
            )

        # P0 Day 4 — immediate output registration. Same pattern as
        # CSV/Excel sinks; see _register_output_in_storage_index.
        _register_output_in_storage_index(ctx, file_path, "json")

        return source

    @staticmethod
    def default_params() -> dict[str, Any]:
        return {"file_path": "", "format": "array"}

    @staticmethod
    def param_schema() -> list[dict]:
        return [
            {"name": "file_path", "type": "text", "label": "File Path",
             "placeholder": "output/results.json"},
            {"name": "format", "type": "select", "label": "JSON Format",
             "options": ["array", "lines"], "default": "array",
             "description": "array = JSON array, lines = newline-delimited JSON"},
        ]


# ── Excel Sink ──

@register(StepType.EXCEL_SINK)
class ExcelSinkNode(BaseNode):
    """Write data to an Excel (.xlsx) file using openpyxl."""
    display_name = "Excel Sink"
    category = "output"
    description = "Write data to an Excel spreadsheet"

    def execute(self, ctx: ExecutionContext) -> duckdb.DuckDBPyRelation:
        source = _get_input(ctx, self.params)
        file_path = _resolve_output_path(
            self.params.get("file_path", ""), ctx.data_dir, "output.xlsx"
        )
        sheet_name = self.params.get("sheet_name", "Sheet1")
        freeze_header = self.params.get("freeze_header", True)

        try:
            import openpyxl
        except ImportError:
            raise ValueError(
                "Excel Sink: openpyxl is required. Install it: pip install openpyxl"
            )

        # Extract data from relation
        ctx.conn.register("__excel_sink_data", source)
        columns = source.columns
        rows = ctx.conn.sql("SELECT * FROM __excel_sink_data").fetchall()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name

        # Write header
        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = openpyxl.styles.Font(bold=True)

        # Write data rows
        for row_idx, row in enumerate(rows, 2):
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Auto-size columns (approximate)
        for col_idx, col_name in enumerate(columns, 1):
            max_len = len(str(col_name))
            for row in rows[:100]:  # Sample first 100 rows for width
                val_len = len(str(row[col_idx - 1])) if row[col_idx - 1] is not None else 0
                max_len = max(max_len, val_len)
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_len + 2, 50)

        # Freeze header row
        if freeze_header:
            ws.freeze_panes = "A2"

        wb.save(file_path)
        wb.close()

        # P0 Day 4 — register the .xlsx in the storage index immediately.
        _register_output_in_storage_index(ctx, file_path, "xlsx")

        return source

    @staticmethod
    def default_params() -> dict[str, Any]:
        return {"file_path": "", "sheet_name": "Sheet1", "freeze_header": True}

    @staticmethod
    def param_schema() -> list[dict]:
        return [
            {"name": "file_path", "type": "text", "label": "File Path",
             "placeholder": "output/report.xlsx"},
            {"name": "sheet_name", "type": "text", "label": "Sheet Name", "default": "Sheet1"},
            {"name": "freeze_header", "type": "boolean", "label": "Freeze Header Row",
             "default": True},
        ]


# ── S3/MinIO Sink ──

@register(StepType.S3_SINK)
class S3SinkNode(BaseNode):
    """Upload data to S3 or MinIO object storage."""
    display_name = "S3 / MinIO Sink"
    category = "output"
    description = "Upload data to S3 or MinIO object storage"

    def execute(self, ctx: ExecutionContext) -> duckdb.DuckDBPyRelation:
        source = _get_input(ctx, self.params)

        bucket = self.params.get("bucket", "")
        key = self.params.get("key", "")
        endpoint = self.params.get("endpoint", "")
        access_key = self.params.get("access_key", "")
        secret_key = self.params.get("secret_key", "")
        region = self.params.get("region", "us-east-1")
        file_format = self.params.get("format", "parquet")
        connection_id = self.params.get("connection_id", "")

        # Load connection config
        if connection_id:
            from fpulse.nodes.db_source import _get_connection_config
            result = _get_connection_config(connection_id)
            if result:
                config, _ = result
                endpoint = config.get("endpoint", endpoint)
                access_key = config.get("access_key", access_key)
                secret_key = config.get("secret_key", secret_key)
                region = config.get("region", region)
                bucket = config.get("bucket", bucket)

        if not bucket or not key:
            raise ValueError("S3 Sink: bucket and key are required")

        # Write to temp file first
        suffix_map = {"parquet": ".parquet", "csv": ".csv", "json": ".json"}
        suffix = suffix_map.get(file_format, ".parquet")

        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
            tmp_path = tmp.name

        try:
            ctx.conn.register("__s3_sink_data", source)
            if file_format == "csv":
                ctx.conn.sql(f"COPY __s3_sink_data TO '{tmp_path}' (FORMAT CSV, HEADER)")
            elif file_format == "json":
                ctx.conn.sql(f"COPY __s3_sink_data TO '{tmp_path}' (FORMAT JSON, ARRAY true)")
            else:
                ctx.conn.sql(f"COPY __s3_sink_data TO '{tmp_path}' (FORMAT PARQUET)")

            # Upload to S3
            self._upload(tmp_path, bucket, key, endpoint, access_key, secret_key, region)
        finally:
            try:
                os.unlink(tmp_path)
            except OSError:
                pass

        return source

    @staticmethod
    def _upload(local_path: str, bucket: str, key: str, endpoint: str,
                access_key: str, secret_key: str, region: str):
        """Upload file to S3/MinIO."""
        try:
            import boto3
        except ImportError:
            raise ValueError(
                "S3 Sink: boto3 is required for upload. Install it: pip install boto3"
            )

        client_kwargs: dict[str, Any] = {"region_name": region}
        if endpoint:
            client_kwargs["endpoint_url"] = endpoint
        if access_key and secret_key:
            client_kwargs["aws_access_key_id"] = access_key
            client_kwargs["aws_secret_access_key"] = secret_key

        s3 = boto3.client("s3", **client_kwargs)
        s3.upload_file(local_path, bucket, key)

    @staticmethod
    def default_params() -> dict[str, Any]:
        return {
            "connection_id": "", "bucket": "", "key": "", "endpoint": "",
            "access_key": "", "secret_key": "", "region": "us-east-1", "format": "parquet",
        }

    @staticmethod
    def param_schema() -> list[dict]:
        return [
            {"name": "connection_id", "type": "connection_picker", "label": "Connection",
             "description": "Select a saved S3/MinIO connection, or fill fields below."},
            {"name": "bucket", "type": "text", "label": "Bucket", "required": True,
             "placeholder": "my-output-bucket"},
            {"name": "key", "type": "text", "label": "Object Key", "required": True,
             "placeholder": "processed/2024/results.parquet"},
            {"name": "endpoint", "type": "text", "label": "Endpoint URL",
             "placeholder": "http://localhost:9000",
             "description": "Required for MinIO. Leave empty for AWS S3."},
            {"name": "access_key", "type": "text", "label": "Access Key"},
            {"name": "secret_key", "type": "password", "label": "Secret Key"},
            {"name": "region", "type": "text", "label": "Region", "default": "us-east-1"},
            {"name": "format", "type": "select", "label": "File Format",
             "options": ["parquet", "csv", "json"], "default": "parquet"},
        ]


# ── FTP / SFTP Sink ──

@register(StepType.FTP_SINK)
class FtpSinkNode(BaseNode):
    """Upload the upstream rows as a file to an FTP, FTPS, or SFTP server.

    Writes the relation to a temp file (csv/json/parquet via DuckDB COPY) then
    uploads it. FTP/FTPS use ``ftplib`` (stdlib); SFTP uses ``paramiko``
    (lazy-imported, clear install hint). The remote directory must already
    exist. Passes the input through unchanged (like every other sink)."""
    display_name = "FTP / SFTP Sink"
    category = "output"
    description = "Upload data as a file to an FTP, FTPS, or SFTP server"

    def execute(self, ctx: ExecutionContext) -> duckdb.DuckDBPyRelation:
        source = _get_input(ctx, self.params)

        host = self.params.get("host", "")
        port = self.params.get("port")
        username = self.params.get("username", "anonymous")
        password = self.params.get("password", "")
        private_key = self.params.get("private_key", "")
        remote_path = self.params.get("remote_path", "")
        file_format = self.params.get("format", "csv")
        connection_id = self.params.get("connection_id", "")

        # Reuse the source node's protocol resolver so ftp/ftps/sftp + the
        # generic connector_type inference behave identically on read + write.
        from fpulse.nodes.sources import FtpSourceNode
        protocol = FtpSourceNode._resolve_protocol(self.params)

        host_key_config: dict[str, Any] = dict(self.params)
        if connection_id:
            from fpulse.nodes.db_source import _get_connection_config
            result = _get_connection_config(connection_id)
            if result:
                config, _ = result
                host_key_config.update(config)
                host = config.get("host", host)
                port = config.get("port", port)
                username = config.get("username", username)
                password = config.get("password", password)
                private_key = config.get("private_key", private_key)
                if config.get("protocol"):
                    protocol = str(config["protocol"]).strip().lower()

        if not host:
            raise ValueError("FTP/SFTP Sink: host is required")
        if not remote_path:
            raise ValueError("FTP/SFTP Sink: remote_path is required")

        try:
            eff_port = int(port) if port not in (None, "", 0) else 0
        except (TypeError, ValueError):
            eff_port = 0
        if eff_port == 0:
            eff_port = 22 if protocol == "sftp" else 21
        elif protocol == "sftp" and eff_port == 21:
            eff_port = 22

        suffix_map = {"parquet": ".parquet", "csv": ".csv", "json": ".json"}
        suffix = suffix_map.get(file_format, ".csv")
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
            tmp_path = tmp.name

        try:
            ctx.conn.register("__ftp_sink_data", source)
            if file_format == "parquet":
                ctx.conn.sql(f"COPY __ftp_sink_data TO '{tmp_path}' (FORMAT PARQUET)")
            elif file_format == "json":
                ctx.conn.sql(f"COPY __ftp_sink_data TO '{tmp_path}' (FORMAT JSON, ARRAY true)")
            else:
                ctx.conn.sql(f"COPY __ftp_sink_data TO '{tmp_path}' (FORMAT CSV, HEADER)")

            if protocol == "sftp":
                try:
                    import paramiko
                except ImportError as exc:
                    raise ValueError(
                        "SFTP requires the 'paramiko' package — install it with "
                        "`pip install paramiko`, then retry."
                    ) from exc
                from fpulse.nodes._sftp import build_ssh_client
                ssh = build_ssh_client(host_key_config)
                ckwargs: dict[str, Any] = {
                    "hostname": host, "port": eff_port, "username": username, "timeout": 30,
                }
                pkey = None
                if private_key:
                    from io import StringIO
                    for kcls in (paramiko.RSAKey, paramiko.Ed25519Key, paramiko.ECDSAKey):
                        try:
                            pkey = kcls.from_private_key(StringIO(str(private_key)))
                            break
                        except Exception:  # noqa: BLE001
                            continue
                if pkey is not None:
                    ckwargs["pkey"] = pkey
                elif password:
                    ckwargs["password"] = password
                ssh.connect(**ckwargs)
                try:
                    ssh.open_sftp().put(tmp_path, remote_path)
                finally:
                    ssh.close()
            else:
                import ftplib
                ftp_class = ftplib.FTP_TLS if protocol == "ftps" else ftplib.FTP
                ftp = ftp_class()
                ftp.connect(host, eff_port, timeout=30)
                ftp.login(username, password)
                if protocol == "ftps" and isinstance(ftp, ftplib.FTP_TLS):
                    ftp.prot_p()
                with open(tmp_path, "rb") as fh:
                    ftp.storbinary(f"STOR {remote_path}", fh)
                ftp.quit()
        finally:
            try:
                os.unlink(tmp_path)
            except OSError:
                pass

        return source

    @staticmethod
    def default_params() -> dict[str, Any]:
        return {
            "connection_id": "", "host": "", "protocol": "ftp", "port": 21,
            "username": "anonymous", "password": "", "private_key": "",
            "remote_path": "", "format": "csv",
        }

    @staticmethod
    def param_schema() -> list[dict]:
        return [
            {"name": "connection_id", "type": "connection_picker", "label": "Connection",
             "description": "Select a saved FTP/SFTP connection, or fill fields below."},
            {"name": "host", "type": "text", "label": "Host", "required": True,
             "placeholder": "sftp.example.com"},
            {"name": "protocol", "type": "select", "label": "Protocol",
             "options": ["ftp", "ftps", "sftp"], "default": "ftp",
             "description": "sftp = SSH File Transfer (needs paramiko); ftps = FTP over TLS."},
            {"name": "port", "type": "number", "label": "Port", "default": 21,
             "description": "FTP/FTPS = 21, SFTP = 22 (auto-corrected if left at 21)."},
            {"name": "username", "type": "text", "label": "Username", "default": "anonymous"},
            {"name": "password", "type": "password", "label": "Password"},
            {"name": "private_key", "type": "password", "label": "Private Key (SFTP, optional)",
             "description": "PEM private key for SFTP key-based auth — alternative to password."},
            {"name": "remote_path", "type": "text", "label": "Remote File Path", "required": True,
             "placeholder": "/uploads/results.csv  (remote directory must exist)"},
            {"name": "format", "type": "select", "label": "File Format",
             "options": ["csv", "json", "parquet"], "default": "csv"},
        ]


# ── Kafka Sink ──

@register(StepType.KAFKA_SINK)
class KafkaSinkNode(BaseNode):
    """Send data to a Kafka topic as JSON messages."""
    display_name = "Kafka Sink"
    category = "output"
    description = "Send data to a Kafka topic"

    def execute(self, ctx: ExecutionContext) -> duckdb.DuckDBPyRelation:
        source = _get_input(ctx, self.params)

        topic = self.params.get("topic", "")
        bootstrap_servers = self.params.get("bootstrap_servers", "localhost:9092")
        key_column = self.params.get("key_column", "")
        batch_size = self.params.get("batch_size", 500)
        connection_id = self.params.get("connection_id", "")

        if not topic:
            raise ValueError("Kafka Sink: topic is required")

        if connection_id:
            from fpulse.nodes.db_source import _get_connection_config
            result = _get_connection_config(connection_id)
            if result:
                config, _ = result
                bootstrap_servers = config.get("bootstrap_servers", bootstrap_servers)

        # Convert relation to list of dicts
        records = _relation_to_dicts(ctx.conn, source, "__kafka_sink_export")

        # Try confluent_kafka first
        try:
            self._send_confluent(records, topic, bootstrap_servers, key_column, batch_size)
            return source
        except ImportError:
            pass

        # Try kafka-python
        try:
            self._send_kafka_python(records, topic, bootstrap_servers, key_column)
            return source
        except ImportError:
            raise ValueError(
                "Kafka Sink: no Kafka client library installed. "
                "Install one: pip install confluent-kafka  OR  pip install kafka-python"
            )

    @staticmethod
    def _send_confluent(records: list[dict], topic: str, bootstrap_servers: str,
                         key_column: str, batch_size: int):
        """Send messages using confluent_kafka."""
        from confluent_kafka import Producer  # type: ignore

        producer = Producer({"bootstrap.servers": bootstrap_servers})
        count = 0

        for record in records:
            key_bytes = None
            if key_column and key_column in record:
                key_bytes = str(record[key_column]).encode("utf-8")

            value_bytes = json.dumps(record, default=str).encode("utf-8")
            producer.produce(topic, value=value_bytes, key=key_bytes)
            count += 1

            if count % batch_size == 0:
                producer.flush()

        producer.flush()

    @staticmethod
    def _send_kafka_python(records: list[dict], topic: str, bootstrap_servers: str,
                            key_column: str):
        """Send messages using kafka-python."""
        from kafka import KafkaProducer  # type: ignore

        producer = KafkaProducer(
            bootstrap_servers=bootstrap_servers.split(","),
            value_serializer=lambda v: json.dumps(v, default=str).encode("utf-8"),
            key_serializer=lambda k: k.encode("utf-8") if k else None,
        )

        for record in records:
            key = str(record.get(key_column, "")) if key_column else None
            producer.send(topic, value=record, key=key)

        producer.flush()
        producer.close()

    @staticmethod
    def default_params() -> dict[str, Any]:
        return {
            "connection_id": "", "topic": "", "bootstrap_servers": "localhost:9092",
            "key_column": "", "batch_size": 500,
        }

    @staticmethod
    def param_schema() -> list[dict]:
        return [
            {"name": "connection_id", "type": "connection_picker", "label": "Connection",
             "description": "Select a saved Kafka connection, or fill fields below."},
            {"name": "topic", "type": "text", "label": "Topic", "required": True,
             "placeholder": "events.processed"},
            {"name": "bootstrap_servers", "type": "text", "label": "Bootstrap Servers",
             "default": "localhost:9092"},
            {"name": "key_column", "type": "text", "label": "Key Column",
             "description": "Column to use as Kafka message key. Leave empty for no key."},
            {"name": "batch_size", "type": "number", "label": "Batch Size", "default": 500,
             "description": "Number of messages between producer flushes."},
        ]


# ── API Sink ──

@register(StepType.API_SINK)
class ApiSinkNode(BaseNode):
    """POST data to a REST API endpoint."""
    display_name = "API Sink"
    category = "output"
    description = "POST data to a REST API endpoint"

    def execute(self, ctx: ExecutionContext) -> duckdb.DuckDBPyRelation:
        source = _get_input(ctx, self.params)

        url = self.params.get("url", "")
        method = self.params.get("method", "POST").upper()
        headers_extra = self.params.get("headers", {})
        batch_mode = self.params.get("batch_mode", "bulk")
        connection_id = self.params.get("connection_id", "")

        if connection_id:
            from fpulse.nodes.db_source import _get_connection_config
            result = _get_connection_config(connection_id)
            if result:
                config, _ = result
                base_url = config.get("base_url", "").rstrip("/")
                path = self.params.get("path", "")
                if path:
                    url = f"{base_url}{path}" if path.startswith("/") else f"{base_url}/{path}"
                elif not url:
                    url = base_url
                if config.get("api_key"):
                    headers_extra.setdefault("Authorization", f"Bearer {config['api_key']}")
                elif config.get("token"):
                    headers_extra.setdefault("Authorization", f"Bearer {config['token']}")

        if not url:
            raise ValueError("API Sink: URL is required")

        records = _relation_to_dicts(ctx.conn, source, "__api_sink_export")

        headers = {
            "Content-Type": "application/json",
            "User-Agent": "F-Pulse/0.6.0",
        }
        headers.update(headers_extra)

        if batch_mode == "bulk":
            # Send all records in a single request
            body = json.dumps(records, default=str).encode("utf-8")
            self._send_request(url, method, headers, body)
        else:
            # Send one request per record
            for record in records:
                body = json.dumps(record, default=str).encode("utf-8")
                self._send_request(url, method, headers, body)

        return source

    def _send_request(self, url: str, method: str, headers: dict, body: bytes) -> dict:
        """Send HTTP request with retry and rate limit handling. Returns response info."""
        import time as _time
        max_retries = int(self.params.get("max_retries", 3))
        retry_delay = float(self.params.get("retry_delay", 1))
        timeout = int(self.params.get("timeout", 30))

        # API Sink POSTs the dataset out — validate the target, then refuse
        # redirects so a 3xx can't bounce it into an internal host after the
        # check (SSRF / data-exfiltration guard).
        try:
            _ssrf_check_url(url, allow_private_env=_HTTP_ALLOW_PRIVATE_ENV)
        except _SsrfBlockedError as exc:
            raise ValueError(f"API Sink: URL blocked by SSRF policy: {exc}") from exc

        opener = _no_redirect_opener()
        last_error = None
        for attempt in range(max_retries + 1):
            try:
                req = urllib.request.Request(url, method=method, headers=headers, data=body)
                with opener.open(req, timeout=timeout) as resp:
                    resp_body = resp.read().decode("utf-8", errors="replace")
                    return {"status": resp.status, "body": resp_body}
            except urllib.error.HTTPError as exc:
                last_error = exc
                if exc.code == 429 or exc.code >= 500:
                    if attempt < max_retries:
                        retry_after = exc.headers.get("Retry-After")
                        wait = float(retry_after) if retry_after else retry_delay * (2 ** attempt)
                        _time.sleep(min(wait, 60))
                        continue
                raise ValueError(f"API Sink: HTTP {exc.code} from {url}: {exc.reason}") from exc
            except urllib.error.URLError as exc:
                last_error = exc
                if attempt < max_retries:
                    _time.sleep(retry_delay * (2 ** attempt))
                    continue
                raise ValueError(f"API Sink: cannot reach {url}: {exc.reason}") from exc

        raise ValueError(f"API Sink: failed after {max_retries} retries: {last_error}")

    @staticmethod
    def default_params() -> dict[str, Any]:
        return {
            "connection_id": "", "url": "", "path": "", "method": "POST",
            "headers": {}, "batch_mode": "bulk",
            "max_retries": 3, "retry_delay": 1, "timeout": 30,
            "on_error": "fail",
        }

    @staticmethod
    def param_schema() -> list[dict]:
        return [
            {"name": "connection_id", "type": "connection_picker", "label": "Connection",
             "tab": "Target",
             "description": "Select a saved REST API connection, or use URL below."},
            {"name": "url", "type": "text", "label": "URL",
             "tab": "Target",
             "placeholder": "https://api.example.com/ingest"},
            {"name": "path", "type": "text", "label": "API Path",
             "tab": "Target",
             "placeholder": "/v1/data",
             "description": "Appended to the connection's base_url."},
            {"name": "method", "type": "select", "label": "HTTP Method",
             "options": ["POST", "PUT", "PATCH"], "default": "POST",
             "tab": "Target"},
            {"name": "headers", "type": "key_value", "label": "Extra Headers",
             "tab": "Target"},
            {"name": "batch_mode", "type": "select", "label": "Batch Mode",
             "options": ["bulk", "per_record"], "default": "bulk",
             "tab": "Behavior",
             "description": "bulk = single request with all records, per_record = one request per row."},
            {"name": "on_error", "type": "select", "label": "On Error",
             "options": ["fail", "continue"], "default": "fail",
             "tab": "Behavior",
             "description": "fail = abort, continue = log error and keep going (per_record mode)."},
            {"name": "max_retries", "type": "number", "label": "Max Retries",
             "default": 3, "tab": "Reliability",
             "description": "Retry on 429/5xx errors with exponential backoff."},
            {"name": "retry_delay", "type": "number", "label": "Retry Delay (seconds)",
             "default": 1, "tab": "Reliability"},
            {"name": "timeout", "type": "number", "label": "Request Timeout (seconds)",
             "default": 30, "tab": "Reliability"},
        ]


# ── Webhook Sink ──

# WebhookSinkNode dropped from palette — POSTing data is covered by
# api_sink and the HTTP Request action. Class kept for legacy pipelines.
class WebhookSinkNode(BaseNode):
    """POST data to a webhook URL."""
    display_name = "Webhook Sink"
    category = "output"
    description = "POST data to a webhook URL"

    def execute(self, ctx: ExecutionContext) -> duckdb.DuckDBPyRelation:
        source = _get_input(ctx, self.params)

        webhook_url = self.params.get("webhook_url", "")
        payload_template = self.params.get("payload_template", "")
        headers_extra = self.params.get("headers", {})
        include_metadata = self.params.get("include_metadata", True)

        if not webhook_url:
            raise ValueError("Webhook Sink: webhook_url is required")

        records = _relation_to_dicts(ctx.conn, source, "__webhook_sink_export")

        # Build payload
        if payload_template:
            # User-defined payload template — substitute {{row_count}} and {{data}}
            payload_str = payload_template.replace(
                "{{row_count}}", str(len(records))
            ).replace(
                "{{data}}", json.dumps(records, default=str)
            )
            body = payload_str.encode("utf-8")
        else:
            payload: dict[str, Any] = {"data": records}
            if include_metadata:
                payload["metadata"] = {
                    "source": "f-pulse",
                    "row_count": len(records),
                    "columns": source.columns,
                }
            body = json.dumps(payload, default=str).encode("utf-8")

        headers = {
            "Content-Type": "application/json",
            "User-Agent": "F-Pulse/0.6.0",
        }
        headers.update(headers_extra)

        # Webhook Sink POSTs the dataset out — same SSRF + no-redirect guard.
        try:
            _ssrf_check_url(webhook_url, allow_private_env=_HTTP_ALLOW_PRIVATE_ENV)
        except _SsrfBlockedError as exc:
            raise ValueError(f"Webhook Sink: URL blocked by SSRF policy: {exc}") from exc

        req = urllib.request.Request(webhook_url, method="POST", headers=headers, data=body)
        try:
            with _no_redirect_opener().open(req, timeout=30):
                pass
        except urllib.error.HTTPError as exc:
            raise ValueError(
                f"Webhook Sink: HTTP {exc.code} from {webhook_url}: {exc.reason}"
            ) from exc
        except urllib.error.URLError as exc:
            raise ValueError(
                f"Webhook Sink: cannot reach {webhook_url}: {exc.reason}"
            ) from exc

        return source

    @staticmethod
    def default_params() -> dict[str, Any]:
        return {
            "webhook_url": "", "payload_template": "", "headers": {},
            "include_metadata": True,
        }

    @staticmethod
    def param_schema() -> list[dict]:
        return [
            {"name": "webhook_url", "type": "text", "label": "Webhook URL", "required": True,
             "placeholder": "https://hooks.slack.com/services/..."},
            {"name": "headers", "type": "key_value", "label": "Extra Headers"},
            {"name": "include_metadata", "type": "boolean", "label": "Include Metadata",
             "default": True,
             "description": "Add row_count, columns, and source info to the payload."},
            {"name": "payload_template", "type": "code", "label": "Custom Payload Template",
             "description": "JSON template with {{data}} and {{row_count}} placeholders. "
                            "Leave empty for default payload structure."},
        ]


# ── Email Sink ──

@register(StepType.EMAIL_SINK)
class EmailSinkNode(BaseNode):
    """Send data via email using SMTP."""
    display_name = "Email Sink"
    category = "output"
    description = "Send data via email (SMTP)"

    def execute(self, ctx: ExecutionContext) -> duckdb.DuckDBPyRelation:
        source = _get_input(ctx, self.params)

        smtp_host = self.params.get("smtp_host", "")
        smtp_port = self.params.get("smtp_port", 587)
        smtp_user = self.params.get("smtp_user", "")
        smtp_password = self.params.get("smtp_password", "")
        use_tls = self.params.get("use_tls", True)
        from_addr = self.params.get("from_address", "")
        to_addrs = self.params.get("to_addresses", [])
        subject = self.params.get("subject", "F-Pulse Pipeline Output")
        body_text = self.params.get("body_text", "")
        attach_format = self.params.get("attach_format", "csv")
        connection_id = self.params.get("connection_id", "")

        if connection_id:
            from fpulse.nodes.db_source import _get_connection_config
            result = _get_connection_config(connection_id)
            if result:
                config, _ = result
                smtp_host = config.get("host", smtp_host)
                smtp_port = int(config.get("port", smtp_port))
                smtp_user = config.get("username", smtp_user)
                smtp_password = config.get("password", smtp_password)
                use_tls = config.get("use_tls", use_tls)
                from_addr = config.get("from_address", from_addr)

        if not smtp_host:
            raise ValueError("Email Sink: SMTP host is required")
        if not to_addrs:
            raise ValueError("Email Sink: at least one recipient address is required")
        if isinstance(to_addrs, str):
            to_addrs = [addr.strip() for addr in to_addrs.split(",") if addr.strip()]

        # Build attachment
        suffix_map = {"csv": ".csv", "json": ".json", "parquet": ".parquet"}
        suffix = suffix_map.get(attach_format, ".csv")

        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
            tmp_path = tmp.name

        try:
            ctx.conn.register("__email_sink_data", source)
            if attach_format == "json":
                ctx.conn.sql(f"COPY __email_sink_data TO '{tmp_path}' (FORMAT JSON, ARRAY true)")
                mime_type = "application/json"
            elif attach_format == "parquet":
                ctx.conn.sql(f"COPY __email_sink_data TO '{tmp_path}' (FORMAT PARQUET)")
                mime_type = "application/octet-stream"
            else:
                ctx.conn.sql(f"COPY __email_sink_data TO '{tmp_path}' (FORMAT CSV, HEADER)")
                mime_type = "text/csv"

            # Build email
            msg = EmailMessage()
            msg["Subject"] = subject
            msg["From"] = from_addr or smtp_user
            msg["To"] = ", ".join(to_addrs)

            row_count = len(ctx.conn.sql("SELECT * FROM __email_sink_data").fetchall())
            columns = source.columns

            if body_text:
                msg.set_content(body_text)
            else:
                msg.set_content(
                    f"F-Pulse Pipeline Output\n\n"
                    f"Rows: {row_count}\n"
                    f"Columns: {', '.join(columns)}\n\n"
                    f"Data is attached as {attach_format.upper()}."
                )

            # Attach file
            with open(tmp_path, "rb") as f:
                attachment_data = f.read()

            maintype, subtype = mime_type.split("/")
            msg.add_attachment(
                attachment_data,
                maintype=maintype,
                subtype=subtype,
                filename=f"pipeline_output{suffix}",
            )

            # Send
            if use_tls:
                server = smtplib.SMTP(smtp_host, int(smtp_port), timeout=30)
                server.starttls()
            else:
                server = smtplib.SMTP(smtp_host, int(smtp_port), timeout=30)

            if smtp_user and smtp_password:
                server.login(smtp_user, smtp_password)

            server.send_message(msg)
            server.quit()
        finally:
            try:
                os.unlink(tmp_path)
            except OSError:
                pass

        return source

    @staticmethod
    def default_params() -> dict[str, Any]:
        return {
            "connection_id": "", "smtp_host": "", "smtp_port": 587,
            "smtp_user": "", "smtp_password": "", "use_tls": True,
            "from_address": "", "to_addresses": [], "subject": "F-Pulse Pipeline Output",
            "body_text": "", "attach_format": "csv",
        }

    @staticmethod
    def param_schema() -> list[dict]:
        return [
            {"name": "connection_id", "type": "connection_picker", "label": "Connection",
             "description": "Select a saved SMTP connection, or fill fields below."},
            {"name": "smtp_host", "type": "text", "label": "SMTP Host", "required": True,
             "placeholder": "smtp.gmail.com"},
            {"name": "smtp_port", "type": "number", "label": "SMTP Port", "default": 587},
            {"name": "smtp_user", "type": "text", "label": "Username",
             "placeholder": "user@example.com"},
            {"name": "smtp_password", "type": "password", "label": "Password"},
            {"name": "use_tls", "type": "boolean", "label": "Use TLS", "default": True},
            {"name": "from_address", "type": "text", "label": "From Address",
             "placeholder": "pipeline@example.com",
             "description": "Defaults to SMTP username if not set."},
            {"name": "to_addresses", "type": "text", "label": "To Addresses", "required": True,
             "placeholder": "user@example.com, team@example.com",
             "description": "Comma-separated email addresses."},
            {"name": "subject", "type": "text", "label": "Subject",
             "default": "F-Pulse Pipeline Output"},
            {"name": "body_text", "type": "code", "label": "Email Body",
             "description": "Plain text email body. Leave empty for auto-generated summary."},
            {"name": "attach_format", "type": "select", "label": "Attachment Format",
             "options": ["csv", "json", "parquet"], "default": "csv"},
        ]


# ── Delta Sink ──

@register(StepType.DELTA_SINK)
class DeltaSinkNode(BaseNode):
    """Write data as Parquet files in a Delta-compatible layout."""
    display_name = "Delta Lake Sink"
    category = "output"
    description = "Write data as a Delta Lake table"

    def execute(self, ctx: ExecutionContext) -> duckdb.DuckDBPyRelation:
        source = _get_input(ctx, self.params)

        table_path = self.params.get("table_path", "")
        if not table_path:
            raise ValueError("Delta Sink: table_path is required")

        if not os.path.isabs(table_path):
            table_path = os.path.join(ctx.data_dir, table_path)

        mode = self.params.get("mode", "overwrite")
        partition_by = self.params.get("partition_by", [])

        # Try deltalake library first
        try:
            import deltalake  # type: ignore
            self._write_with_deltalake(ctx, source, table_path, mode, partition_by)
            return source
        except ImportError:
            pass

        # Fallback: write as plain Parquet in Delta-like directory structure
        self._write_parquet_fallback(ctx, source, table_path, mode, partition_by)
        return source

    @staticmethod
    def _write_with_deltalake(ctx: ExecutionContext, source: duckdb.DuckDBPyRelation,
                               table_path: str, mode: str,
                               partition_by: list[str]):
        """Write using the deltalake Python library."""
        import deltalake  # type: ignore

        ctx.conn.register("__delta_sink_data", source)
        arrow_table = ctx.conn.sql("SELECT * FROM __delta_sink_data").arrow()

        write_kwargs: dict[str, Any] = {"mode": mode}
        if partition_by:
            write_kwargs["partition_by"] = partition_by

        deltalake.write_deltalake(table_path, arrow_table, **write_kwargs)

    @staticmethod
    def _write_parquet_fallback(ctx: ExecutionContext, source: duckdb.DuckDBPyRelation,
                                 table_path: str, mode: str,
                                 partition_by: list[str]):
        """Fallback: write as Parquet files in a Delta-like directory."""
        os.makedirs(table_path, exist_ok=True)

        ctx.conn.register("__delta_sink_data", source)

        if mode == "overwrite":
            # Remove existing Parquet files
            import glob as glob_mod
            for old_file in glob_mod.glob(os.path.join(table_path, "*.parquet")):
                os.remove(old_file)

        # Generate a unique filename
        import uuid
        file_name = f"part-{uuid.uuid4().hex[:12]}.parquet"
        file_path = os.path.join(table_path, file_name)

        if partition_by:
            # Use DuckDB's partitioned COPY
            part_cols = ", ".join(f'"{c}"' for c in partition_by)
            ctx.conn.sql(
                f"COPY __delta_sink_data TO '{table_path}' "
                f"(FORMAT PARQUET, PARTITION_BY ({part_cols}))"
            )
        else:
            ctx.conn.sql(
                f"COPY __delta_sink_data TO '{file_path}' (FORMAT PARQUET)"
            )

    @staticmethod
    def default_params() -> dict[str, Any]:
        return {"table_path": "", "mode": "overwrite", "partition_by": []}

    @staticmethod
    def param_schema() -> list[dict]:
        return [
            {"name": "table_path", "type": "text", "label": "Table Path", "required": True,
             "placeholder": "lakehouse/silver/customers",
             "description": "Directory path for the Delta Lake table."},
            {"name": "mode", "type": "select", "label": "Write Mode",
             "options": ["overwrite", "append"], "default": "overwrite",
             "description": "overwrite = replace table, append = add to existing data"},
            {"name": "partition_by", "type": "column_list", "label": "Partition By",
             "description": "Columns to partition the data by (e.g. date, region)."},
        ]


# ── Warehouse Sink ──

@register(StepType.WAREHOUSE_SINK)
class WarehouseSinkNode(BaseNode):
    """Write data to a data warehouse with schema evolution support.

    Wraps DB_SINK with automatic CREATE TABLE or ALTER TABLE to handle
    new columns gracefully.
    """
    display_name = "Warehouse Sink"
    category = "output"
    description = "Write to a data warehouse with schema evolution"

    def execute(self, ctx: ExecutionContext) -> duckdb.DuckDBPyRelation:
        source = _get_input(ctx, self.params)
        # Apply Mapping-tab settings (rename + skip) to the source
        # relation BEFORE any schema-evolution / table-create logic
        # below runs against its columns. No-op when the Mapping tab
        # hasn't been touched.
        from fpulse.nodes._column_mapping import apply_column_mapping
        source = apply_column_mapping(source, self.params)

        table = self.params.get("table", "")
        schema_name = self.params.get("schema", "public")
        mode = self.params.get("mode", "create")
        connection_id = self.params.get("connection_id", "")
        auto_evolve = self.params.get("auto_evolve", True)

        if not table:
            raise ValueError("Warehouse Sink: table name is required")

        # If no connection, use DuckDB in-memory with schema evolution
        if not connection_id:
            return self._write_duckdb(ctx, source, table, schema_name, mode, auto_evolve)

        # Real database — delegate to DB Sink logic with schema evolution
        return self._write_real(ctx, source, table, schema_name, mode,
                                connection_id, auto_evolve)

    def _write_duckdb(self, ctx: ExecutionContext, source: duckdb.DuckDBPyRelation,
                       table: str, schema_name: str, mode: str,
                       auto_evolve: bool) -> duckdb.DuckDBPyRelation:
        """Write to DuckDB in-memory with schema-policy-driven evolution."""
        ctx.conn.register("__wh_sink_data", source)
        full_table = f'"{schema_name}"."{table}"' if schema_name != "public" else f'"{table}"'

        pre_sql = (self.params.get("pre_sql") or "").strip()
        post_sql = (self.params.get("post_sql") or "").strip()
        policy_value = _resolve_schema_policy(self.params)
        display_name = f"{schema_name}.{table}" if schema_name else table

        if pre_sql:
            ctx.conn.sql(pre_sql)

        if mode == "create":
            # Explicit "drop and recreate" — policy doesn't gate this
            # path (the user opted into wholesale replacement). We
            # still emit an event when the previous shape differed so
            # the audit trail captures the recreate.
            ctx.conn.sql(f"CREATE OR REPLACE TABLE {full_table} AS SELECT * FROM __wh_sink_data")
        elif mode == "append":
            # Read the existing column shape from DuckDB's information_schema.
            # Tuple is (name, data_type, is_nullable) — keep all three so the
            # policy can see nullability and type drift, not just adds/drops.
            try:
                existing_rows = ctx.conn.sql(
                    "SELECT column_name, data_type, is_nullable FROM information_schema.columns "
                    f"WHERE table_name = '{table}'"
                ).fetchall()
            except Exception:
                existing_rows = []

            existing_cols = [
                {"name": r[0], "type": r[1], "nullable": str(r[2]).upper() != "NO"}
                for r in existing_rows
            ]
            incoming_cols = [
                {"name": c, "type": str(t), "nullable": True}
                for c, t in zip(source.columns, source.types)
            ]

            if not existing_cols:
                # Table doesn't exist — first write, policy short-circuits to ok.
                ctx.conn.sql(
                    f"CREATE TABLE {full_table} AS SELECT * FROM __wh_sink_data"
                )
            else:
                decision = evaluate_policy(existing_cols, incoming_cols, policy_value)
                if not decision.ok:
                    _publish_warehouse_drift_event(
                        ctx,
                        step_id=self.params.get("_step_id", ""),
                        table_name=display_name,
                        decision=decision,
                        applied=False,
                        rejection_reason=decision.rejection_reason or "",
                    )
                    decision.raise_if_rejected()

                # Apply additive changes when the policy allows.
                for add in decision.adds:
                    ctx.conn.sql(
                        f'ALTER TABLE {full_table} ADD COLUMN "{add.column}" {add.to_type or "VARCHAR"}'
                    )
                # DuckDB ALTER TABLE doesn't widen types in-place generically;
                # widening is silently safe at INSERT time when DuckDB coerces
                # the narrower source into the wider destination. So the
                # widen + force passes are no-ops here — the INSERT below
                # carries the values across. The event still fires below.

                ctx.conn.sql(f"INSERT INTO {full_table} SELECT * FROM __wh_sink_data")

                if decision.has_drift:
                    _publish_warehouse_drift_event(
                        ctx,
                        step_id=self.params.get("_step_id", ""),
                        table_name=display_name,
                        decision=decision,
                        applied=True,
                    )
        elif mode == "truncate":
            try:
                ctx.conn.sql(f"DELETE FROM {full_table}")
            except Exception:
                ctx.conn.sql(
                    f"CREATE TABLE {full_table} AS SELECT * FROM __wh_sink_data WHERE false"
                )
            ctx.conn.sql(f"INSERT INTO {full_table} SELECT * FROM __wh_sink_data")

        if post_sql:
            ctx.conn.sql(post_sql)

        return source

    def _parse_merge_key(self) -> list[str]:
        """B2.1 — parse the comma-separated merge_key param into a clean
        list of column names. Empty list when unset."""
        raw = (self.params.get("merge_key") or "").strip()
        if not raw:
            return []
        return [c.strip() for c in raw.split(",") if c.strip()]

    @staticmethod
    def _normalize_dialect(conn_type: str) -> str:
        """Map a connection's conn_type onto the bulk-load registry's
        dialect key (the registry uses 'postgresql', not 'postgres')."""
        aliases = {
            "postgres": "postgresql", "postgresql": "postgresql",
            "mssql": "mssql", "sqlserver": "mssql",
            "snowflake": "snowflake", "redshift": "redshift",
            "bigquery": "bigquery", "duckdb": "duckdb",
        }
        c = (conn_type or "").lower()
        return aliases.get(c, c)

    def _write_merge(self, ctx: ExecutionContext, source: duckdb.DuckDBPyRelation,
                      table: str, schema_name: str, conn_type: str,
                      config: dict) -> duckdb.DuckDBPyRelation:
        """B2.1 — upsert via the bulk-load MERGE path. Maps the
        operator's merge_key onto BulkLoadRequest.primary_key so the
        existing per-dialect MERGE / ON CONFLICT generation runs. Returns
        the source relation unchanged (passthrough).

        [LIVE-SMOKE] the actual MERGE execution needs a live warehouse;
        this method's wiring (key mapping + loader selection + request
        construction) is unit-tested with a fake loader."""
        merge_key = self._parse_merge_key()
        if not merge_key:
            raise ValueError(
                "Warehouse Sink: merge mode requires Merge Key Column(s). "
                "Set the natural key column(s) used to match rows for upsert."
            )
        from fpulse.engine.bulk_load import registry as _bl_reg
        from fpulse.engine.bulk_load.types import BulkLoadRequest
        dialect = self._normalize_dialect(conn_type)
        loader = _bl_reg.get(dialect)
        if loader is None or not loader.is_available():
            raise ValueError(
                f"Warehouse Sink: merge mode needs an installed bulk loader "
                f"for '{conn_type}'. Install the driver, or use append mode."
            )
        req = BulkLoadRequest(
            conn_type=dialect,
            config=config,
            table=table,
            schema_name=schema_name or "public",
            mode="merge",
            primary_key=merge_key,
            relation=source,
            duckdb_conn=ctx.conn,
            columns=list(source.columns),
        )
        loader.load(req)
        return source

    def _write_real(self, ctx: ExecutionContext, source: duckdb.DuckDBPyRelation,
                     table: str, schema_name: str, mode: str,
                     connection_id: str, auto_evolve: bool) -> duckdb.DuckDBPyRelation:
        """Write to a real database with schema evolution and proper types."""
        from fpulse.nodes.db_source import _get_connection_config
        result = _get_connection_config(connection_id)
        if not result:
            raise ValueError(f"Warehouse Sink: connection '{connection_id}' not found")

        config, conn_type = result

        # B2.1 (2026-06-08, docs/design/backfill-ux-1.2.md) — merge
        # (upsert) mode delegates to the bulk-load MERGE path, which
        # already implements per-dialect upsert keyed on primary_key
        # (postgres ON CONFLICT, mssql / snowflake MERGE). We map the
        # operator's merge_key onto primary_key here. Handled BEFORE the
        # create/append/truncate connection block so the two write
        # paths' connection lifecycles don't interleave.
        # [LIVE-SMOKE] real-warehouse upsert correctness must be verified
        # against a live DB before relying on it.
        if mode == "merge":
            return self._write_merge(ctx, source, table, schema_name,
                                     conn_type, config)

        batch_size = max(int(self.params.get("batch_size", 1000)), 1)
        pre_sql = (self.params.get("pre_sql") or "").strip()
        post_sql = (self.params.get("post_sql") or "").strip()

        # Extract data
        ctx.conn.register("__wh_sink_export", source)
        columns = source.columns
        types = source.types
        rows = ctx.conn.sql("SELECT * FROM __wh_sink_export").fetchall()

        # Get database connection
        db_conn = self._connect(conn_type, config)
        try:
            cur = db_conn.cursor()
            full_table = f'"{schema_name}"."{table}"' if schema_name else f'"{table}"'

            # Pre-write SQL
            if pre_sql:
                cur.execute(pre_sql)

            if mode == "create":
                # Drop and recreate with proper types
                try:
                    cur.execute(f"DROP TABLE IF EXISTS {full_table}")
                except Exception:
                    pass

                col_defs = self._typed_col_defs(columns, types, conn_type)
                cur.execute(f"CREATE TABLE {full_table} ({col_defs})")
                self._insert_rows(cur, full_table, columns, rows, conn_type,
                                  batch_size)

            elif mode == "append":
                # Policy-driven schema evolution. See schema_policy.py
                # for the four-value enum + acceptance rules.
                policy_value = _resolve_schema_policy(self.params)
                display_name = f"{schema_name}.{table}" if schema_name else table

                existing_cols_raw = self._get_existing_columns(
                    cur, table, schema_name, conn_type
                )
                if existing_cols_raw is None:
                    # Table doesn't exist — first write. Policy short-
                    # circuits to ok and we create with proper types.
                    col_defs = self._typed_col_defs(columns, types, conn_type)
                    cur.execute(f"CREATE TABLE {full_table} ({col_defs})")
                else:
                    # We have the existing column names; the policy
                    # checks adds + drops on this basis. Type drift
                    # detection on a real DB needs information_schema
                    # type lookups, which differ per dialect. For OSS
                    # v1 we keep policy decisions on the name set —
                    # type widening + narrowing are detected via the
                    # incoming relation's DuckDB types matching the
                    # existing names; deeper introspection is a
                    # follow-up once the surface stabilises.
                    existing_for_policy = [
                        {"name": c, "type": "VARCHAR", "nullable": True}
                        for c in existing_cols_raw
                    ]
                    incoming_for_policy = [
                        {"name": c, "type": str(t), "nullable": True}
                        for c, t in zip(columns, types)
                    ]
                    decision = evaluate_policy(
                        existing_for_policy, incoming_for_policy, policy_value,
                    )
                    if not decision.ok:
                        _publish_warehouse_drift_event(
                            ctx,
                            step_id=self.params.get("_step_id", ""),
                            table_name=display_name,
                            decision=decision,
                            applied=False,
                            rejection_reason=decision.rejection_reason or "",
                        )
                        decision.raise_if_rejected()

                    for add in decision.adds:
                        sql_type = self._map_type(add.to_type or "VARCHAR", conn_type)
                        cur.execute(
                            f'ALTER TABLE {full_table} ADD COLUMN "{add.column}" {sql_type}'
                        )

                    if decision.has_drift:
                        _publish_warehouse_drift_event(
                            ctx,
                            step_id=self.params.get("_step_id", ""),
                            table_name=display_name,
                            decision=decision,
                            applied=True,
                        )

                self._insert_rows(cur, full_table, columns, rows, conn_type,
                                  batch_size)

            elif mode == "truncate":
                try:
                    cur.execute(f"DELETE FROM {full_table}")
                except Exception:
                    col_defs = self._typed_col_defs(columns, types, conn_type)
                    cur.execute(f"CREATE TABLE {full_table} ({col_defs})")
                self._insert_rows(cur, full_table, columns, rows, conn_type,
                                  batch_size)

            # Post-write SQL
            if post_sql:
                cur.execute(post_sql)

            db_conn.commit()
        finally:
            db_conn.close()

        return source

    @staticmethod
    def _connect(conn_type: str, config: dict):
        """Create a database connection based on type."""
        host = config.get("host")
        port = config.get("port")
        database = config.get("database")
        user = config.get("user") or config.get("username")
        password = config.get("password")
        schema = config.get("schema")

        if conn_type == "sqlite":
            import sqlite3
            return sqlite3.connect(config.get("database") or config.get("file", ""))
        elif conn_type == "postgresql":
            import psycopg2  # type: ignore
            return psycopg2.connect(
                host=host, port=port or 5432, dbname=database,
                user=user, password=password, connect_timeout=10,
                options=f"-c search_path={schema}" if schema else None,
            )
        elif conn_type == "mysql":
            import pymysql  # type: ignore
            return pymysql.connect(
                host=host, port=int(port or 3306), database=database,
                user=user, password=password, connect_timeout=10,
            )
        elif conn_type == "mssql":
            import pyodbc  # type: ignore
            conn_str = (
                f"DRIVER={{ODBC Driver 17 for SQL Server}};"
                f"SERVER={host},{port or 1433};DATABASE={database};"
                f"UID={user};PWD={password};Connection Timeout=10;"
            )
            return pyodbc.connect(conn_str)
        else:
            raise ValueError(f"Warehouse Sink: unsupported connection type '{conn_type}'")

    @staticmethod
    def _get_existing_columns(cur, table: str, schema_name: str,
                               conn_type: str) -> list[str] | None:
        """Get existing column names for a table, or None if table doesn't exist."""
        try:
            if conn_type == "postgresql":
                cur.execute(
                    "SELECT column_name FROM information_schema.columns "
                    "WHERE table_name = %s AND table_schema = %s",
                    (table, schema_name or "public"),
                )
            elif conn_type == "mysql":
                cur.execute(
                    "SELECT column_name FROM information_schema.columns "
                    "WHERE table_name = %s AND table_schema = DATABASE()",
                    (table,),
                )
            elif conn_type == "mssql":
                cur.execute(
                    "SELECT column_name FROM information_schema.columns "
                    "WHERE table_name = ? AND table_schema = ?",
                    (table, schema_name or "dbo"),
                )
            elif conn_type == "sqlite":
                cur.execute(f'PRAGMA table_info("{table}")')
                rows = cur.fetchall()
                if not rows:
                    return None
                return [row[1] for row in rows]
            else:
                return None

            rows = cur.fetchall()
            return [row[0] for row in rows] if rows else None
        except Exception:
            return None

    @staticmethod
    def _insert_rows(cur, full_table: str, columns: list[str],
                      rows: list[tuple], conn_type: str,
                      batch_size: int = 1000):
        """Insert rows into the table in batches."""
        if not rows:
            return
        col_names = ", ".join(f'"{c}"' for c in columns)
        placeholder = "?" if conn_type in ("sqlite", "mssql") else "%s"
        placeholders = ", ".join([placeholder] * len(columns))
        sql = f"INSERT INTO {full_table} ({col_names}) VALUES ({placeholders})"

        # Batch insert for performance
        for i in range(0, len(rows), batch_size):
            batch = rows[i:i + batch_size]
            cur.executemany(sql, batch)

    # ── Type inference ──────────────────────────────────────────────
    _DUCKDB_TO_SQL = {
        "BIGINT": "BIGINT", "INTEGER": "INTEGER", "SMALLINT": "SMALLINT",
        "TINYINT": "SMALLINT", "HUGEINT": "BIGINT",
        "DOUBLE": "DOUBLE PRECISION", "FLOAT": "REAL",
        "DECIMAL": "DECIMAL", "BOOLEAN": "BOOLEAN",
        "DATE": "DATE", "TIME": "TIME", "TIMESTAMP": "TIMESTAMP",
        "TIMESTAMP WITH TIME ZONE": "TIMESTAMPTZ",
        "VARCHAR": "TEXT", "BLOB": "BYTEA",
    }

    @classmethod
    def _map_type(cls, duckdb_type: str, conn_type: str) -> str:
        """Map a DuckDB type string to the target database type."""
        dtype = str(duckdb_type).upper()
        # Strip DECIMAL precision for lookup
        base = dtype.split("(")[0].strip()
        mapped = cls._DUCKDB_TO_SQL.get(base, "TEXT")

        # Database-specific overrides
        if conn_type == "mysql":
            if mapped == "BOOLEAN":
                return "TINYINT(1)"
            if mapped == "TEXT":
                return "TEXT"
            if mapped == "BYTEA":
                return "LONGBLOB"
            if mapped == "TIMESTAMPTZ":
                return "DATETIME"
        elif conn_type == "mssql":
            if mapped == "BOOLEAN":
                return "BIT"
            if mapped == "TEXT":
                return "NVARCHAR(MAX)"
            if mapped == "DOUBLE PRECISION":
                return "FLOAT"
            if mapped == "REAL":
                return "REAL"
            if mapped == "BYTEA":
                return "VARBINARY(MAX)"
            if mapped == "TIMESTAMPTZ":
                return "DATETIMEOFFSET"
        elif conn_type == "sqlite":
            # SQLite has flexible typing
            if base in ("BIGINT", "INTEGER", "SMALLINT", "TINYINT", "HUGEINT"):
                return "INTEGER"
            if base in ("DOUBLE", "FLOAT", "DECIMAL"):
                return "REAL"
            return "TEXT"

        return mapped

    def _typed_col_defs(self, columns: list[str], types: list,
                        conn_type: str) -> str:
        """Build column definitions with proper types instead of all TEXT."""
        parts = []
        for col, dtype in zip(columns, types):
            sql_type = self._map_type(str(dtype), conn_type)
            parts.append(f'"{col}" {sql_type}')
        return ", ".join(parts)

    @staticmethod
    def default_params() -> dict[str, Any]:
        return {
            "connection_id": "", "table": "", "schema": "public",
            "mode": "create",
            # auto_evolve kept as a legacy alias for back-compat with
            # pipelines saved before 2026-05-27. _resolve_schema_policy()
            # maps it onto schema_policy when the latter is unset.
            "auto_evolve": True,
            "schema_policy": DEFAULT_POLICY.value,
            "batch_size": 1000,
            "pre_sql": "", "post_sql": "",
        }

    @staticmethod
    def param_schema() -> list[dict]:
        return [
            {"name": "connection_id", "type": "connection_picker", "label": "Connection",
             "tab": "Target",
             "description": "Select a saved database connection. Leave empty for DuckDB in-memory."},
            {"name": "table", "type": "text", "label": "Table Name", "required": True,
             "tab": "Target",
             "placeholder": "fact_sales"},
            {"name": "schema", "type": "text", "label": "Schema", "default": "public",
             "tab": "Target",
             "description": "Database schema (PostgreSQL/MSSQL) or leave as public."},
            {"name": "mode", "type": "select", "label": "Write Mode",
             "options": ["create", "append", "truncate", "merge"], "default": "create",
             "tab": "Target",
             "description": "create = drop & recreate, append = add rows, "
                            "truncate = clear & insert, merge = upsert on the "
                            "merge key (insert new rows, update matching rows)"},
            # B2 (2026-06-08, docs/design/backfill-ux-1.2.md) - merge key
            # for upsert mode. The per-dialect MERGE / ON CONFLICT SQL
            # already exists in engine/bulk_load/dialects (keyed on
            # BulkLoadRequest.primary_key); this field is the operator-
            # facing way to populate that key. Wiring merge_key ->
            # request.primary_key in the sink execute() path is the
            # deferred B2.1 integration.
            {"name": "merge_key", "type": "text", "label": "Merge Key Column(s)",
             "tab": "Target",
             "placeholder": "id   (or: customer_id, order_date)",
             "show_when": {"mode": ["merge"]},
             "description": "Comma-separated natural key column(s) used to match "
                            "rows for upsert. New rows are inserted; rows matching "
                            "the key are updated. Warn: the destination should have "
                            "a unique index on these columns, or an upsert may "
                            "overwrite multiple rows."},
            # B4 (2026-06-08, docs/design/backfill-ux-1.2.md) - tombstone
            # column for soft-delete propagation on merge. When set, rows
            # whose tombstone column is truthy are deleted at the
            # destination instead of upserted (see sinks/tombstone.py for
            # the partition helper). Per-dialect DELETE codegen wire-in is
            # the deferred B4.1 integration.
            {"name": "tombstone_column", "type": "text",
             "label": "Tombstone Column (soft-delete propagation)",
             "tab": "Target",
             "placeholder": "is_deleted",
             "show_when": {"mode": ["merge"]},
             "description": "Optional. Column flagging soft-deletes (e.g. "
                            "is_deleted / deleted_at). On merge, rows with this "
                            "column set are deleted at the destination instead of "
                            "upserted. Requires a merge key. For hard-delete "
                            "tracking use the CDC connector (Plus)."},
            schema_policy_param(),
            {"name": "auto_evolve", "type": "boolean", "label": "Auto Schema Evolution (legacy)",
             "default": True, "tab": "Schema",
             "tier": "optional",
             "description": "Legacy flag; ignored when schema_policy is set. Kept for "
                            "pipelines saved before the policy lever shipped."},
            {"name": "batch_size", "type": "number", "label": "Batch Size",
             "default": 1000, "tab": "Performance",
             "description": "Number of rows per INSERT batch (higher = faster, more memory)."},
            {"name": "pre_sql", "type": "sql", "label": "Pre-Write SQL",
             "tab": "Advanced",
             "placeholder": "DELETE FROM staging.temp WHERE batch_date = CURRENT_DATE",
             "description": "SQL to run before writing data."},
            {"name": "post_sql", "type": "sql", "label": "Post-Write SQL",
             "tab": "Advanced",
             "placeholder": "CALL refresh_materialized_view('mv_sales')",
             "description": "SQL to run after writing data."},
        ]
