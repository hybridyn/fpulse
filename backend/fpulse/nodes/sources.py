"""
Extended source nodes for F-Pulse.

Covers: JSON, Parquet, Excel, XML, S3/MinIO, Kafka, FTP, Google Sheets, Delta Lake.
All nodes return duckdb.DuckDBPyRelation for downstream consumption.
"""

from __future__ import annotations

import io
import json
import os
import tempfile
import urllib.request
import urllib.error
from typing import Any, TYPE_CHECKING

# Stage 2.5b: duckdb only used in type annotations (return types and
# function signatures). Helper bodies use the `conn` argument they're
# passed; node `execute()` methods use ctx.conn. Neither needs the
# runtime module import — `from __future__ import annotations` makes
# every annotation a string at runtime.
if TYPE_CHECKING:
    import duckdb

from fpulse.ir.schema import StepType
from fpulse.nodes.base import BaseNode, ExecutionContext
from fpulse.nodes.registry import register
from fpulse.nodes.guardrails import check_file_size, cap_rows


# ── Helpers ──

def _resolve_path(file_path: str, data_dir: str) -> str:
    """Resolve a READ file path: data_dir first, then project-CWD
    fallback. Sample-pack pipelines pass project-relative paths like
    `samples/free-api-pipelines/data/orders.csv` that would otherwise
    get data_dir-doubled (see `_path_utils.py` rationale, 2026-05-26)."""
    if not file_path:
        raise ValueError("File path is required")
    from fpulse.nodes._path_utils import resolve_input_path
    return resolve_input_path(file_path, data_dir)


def _rows_to_relation(conn: duckdb.DuckDBPyConnection, rows: list[dict],
                       table_name: str = "__source_data") -> duckdb.DuckDBPyRelation:
    """Convert a list of dicts into a DuckDB relation."""
    if not rows:
        return conn.sql("SELECT NULL AS empty WHERE false")

    # Collect all keys preserving order
    all_keys: list[str] = []
    seen: set[str] = set()
    for row in rows:
        for k in row:
            if k not in seen:
                all_keys.append(k)
                seen.add(k)

    def fmt(v):
        if v is None:
            return "NULL"
        if isinstance(v, bool):
            return "true" if v else "false"
        if isinstance(v, (int, float)):
            return str(v)
        if isinstance(v, (dict, list)):
            return "'" + json.dumps(v).replace("'", "''") + "'"
        return "'" + str(v).replace("'", "''") + "'"

    value_rows = []
    for row in rows:
        vals = ", ".join(fmt(row.get(k)) for k in all_keys)
        value_rows.append(f"({vals})")

    values_sql = ", ".join(value_rows)
    # Name columns in the VALUES alias — DuckDB's positional auto-naming
    # shifted from `column0` to `col0` across versions, so the old
    # `column{i} AS "name"` rename raised a Binder Error on newer installs.
    quoted_cols = ", ".join(f'"{k}"' for k in all_keys)
    # Unique staging-table name per call so two same-type row-based sources
    # (e.g. two JSON Sources) in one pipeline don't clobber each other's
    # CREATE OR REPLACE TEMP TABLE — which would otherwise make the first
    # source's returned lazy relation silently read the second's data.
    import uuid as _uuid
    _tbl = f"{table_name}_{_uuid.uuid4().hex[:8]}"
    conn.execute(f"CREATE OR REPLACE TEMP TABLE {_tbl} AS SELECT * FROM (VALUES {values_sql}) AS __vals ({quoted_cols})")
    return conn.sql(f"SELECT * FROM {_tbl}")


# ── JSON Source ──

@register(StepType.JSON_SOURCE)
class JsonSourceNode(BaseNode):
    """Read data from a JSON file using DuckDB's native read_json."""
    display_name = "JSON Source"
    category = "source"
    description = "Read data from a JSON file"

    def execute(self, ctx: ExecutionContext) -> duckdb.DuckDBPyRelation:
        file_path = _resolve_path(self.params.get("file_path", ""), ctx.data_dir)

        if not os.path.isfile(file_path):
            raise ValueError(f"JSON Source: file not found: {file_path}")

        check_file_size(file_path)

        json_format = self.params.get("format", "auto")
        records_path = self.params.get("records_path", "")

        # If user specified a records_path, load manually and extract
        if records_path:
            with open(file_path, "r", encoding="utf-8") as f:
                data = json.load(f)

            # Navigate into nested path (e.g. "data.results")
            for key in records_path.split("."):
                if isinstance(data, dict) and key in data:
                    data = data[key]
                else:
                    raise ValueError(f"JSON Source: key '{key}' not found in JSON structure")

            if not isinstance(data, list):
                data = [data]

            return cap_rows(_rows_to_relation(ctx.conn, data, "__json_source"), label="JSON Source", full_run=ctx.full_run)

        # Use DuckDB native JSON reader
        if json_format == "records":
            rel = ctx.conn.read_json(file_path, format="array")
        elif json_format == "lines":
            rel = ctx.conn.read_json(file_path, format="newline_delimited")
        else:
            # Auto-detect: try array first, then newline-delimited
            rel = ctx.conn.read_json(file_path)
        return cap_rows(rel, label="JSON Source", full_run=ctx.full_run)

    @staticmethod
    def default_params() -> dict[str, Any]:
        return {"file_path": "", "format": "auto", "records_path": ""}

    @staticmethod
    def param_schema() -> list[dict]:
        return [
            {"name": "file_path", "type": "file", "label": "File Path", "required": True,
             "placeholder": "data/input.json"},
            {"name": "format", "type": "select", "label": "JSON Format",
             "options": ["auto", "records", "lines"], "default": "auto",
             "description": "auto = DuckDB auto-detect, records = JSON array, lines = newline-delimited"},
            {"name": "records_path", "type": "text", "label": "Records Path",
             "placeholder": "data.results",
             "description": "Dot-separated path to the array within the JSON (e.g. 'data.items')"},
        ]


# ── Parquet Source ──

@register(StepType.PARQUET_SOURCE)
class ParquetSourceNode(BaseNode):
    """Read data from Parquet files using DuckDB's native reader."""
    display_name = "Parquet Source"
    category = "source"
    description = "Read data from a Parquet file"

    def execute(self, ctx: ExecutionContext) -> duckdb.DuckDBPyRelation:
        file_path = _resolve_path(self.params.get("file_path", ""), ctx.data_dir)

        # Support glob patterns like "data/*.parquet"
        if "*" in file_path:
            return cap_rows(ctx.conn.read_parquet(file_path), label="Parquet Source", full_run=ctx.full_run)

        if not os.path.isfile(file_path):
            raise ValueError(f"Parquet Source: file not found: {file_path}")

        check_file_size(file_path)

        columns = self.params.get("columns", [])
        row_group_filter = self.params.get("row_group_filter", "")

        if columns:
            # Read only specified columns for efficiency
            col_list = ", ".join(f'"{c}"' for c in columns)
            return cap_rows(ctx.conn.sql(
                f"SELECT {col_list} FROM read_parquet('{file_path}')"
            ), label="Parquet Source", full_run=ctx.full_run)

        return cap_rows(ctx.conn.read_parquet(file_path), label="Parquet Source", full_run=ctx.full_run)

    @staticmethod
    def default_params() -> dict[str, Any]:
        return {"file_path": "", "columns": []}

    @staticmethod
    def param_schema() -> list[dict]:
        return [
            {"name": "file_path", "type": "file", "label": "File Path", "required": True,
             "placeholder": "data/input.parquet",
             "description": "Supports glob patterns like data/*.parquet"},
            {"name": "columns", "type": "column_list", "label": "Select Columns",
             "description": "Leave empty to read all columns"},
        ]


# ── Excel Source ──

@register(StepType.EXCEL_SOURCE)
class ExcelSourceNode(BaseNode):
    """Read data from Excel (.xlsx/.xls) files using openpyxl."""
    display_name = "Excel Source"
    category = "source"
    description = "Read data from an Excel spreadsheet"

    def execute(self, ctx: ExecutionContext) -> duckdb.DuckDBPyRelation:
        file_path = _resolve_path(self.params.get("file_path", ""), ctx.data_dir)

        if not os.path.isfile(file_path):
            raise ValueError(f"Excel Source: file not found: {file_path}")

        check_file_size(file_path)

        sheet_name = self.params.get("sheet_name", "")
        header_row = self.params.get("header_row", 1)
        skip_rows = self.params.get("skip_rows", 0)

        # Try openpyxl first (most reliable for .xlsx)
        try:
            import openpyxl
            return self._read_with_openpyxl(ctx.conn, file_path, sheet_name,
                                             header_row, skip_rows)
        except ImportError:
            pass

        # Fallback: try DuckDB spatial extension (st_read)
        try:
            ctx.conn.execute("INSTALL spatial; LOAD spatial;")
            layer = sheet_name if sheet_name else ""
            if layer:
                return ctx.conn.sql(
                    f"SELECT * FROM st_read('{file_path}', layer='{layer}')"
                )
            return ctx.conn.sql(f"SELECT * FROM st_read('{file_path}')")
        except Exception:
            raise ValueError(
                "Excel Source: openpyxl is not installed and DuckDB spatial extension "
                "failed. Install openpyxl: pip install openpyxl"
            )

    @staticmethod
    def _read_with_openpyxl(conn: duckdb.DuckDBPyConnection, file_path: str,
                             sheet_name: str, header_row: int,
                             skip_rows: int) -> duckdb.DuckDBPyRelation:
        """Read Excel file using openpyxl and load into DuckDB."""
        import openpyxl

        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

        if sheet_name:
            if sheet_name not in wb.sheetnames:
                wb.close()
                raise ValueError(
                    f"Excel Source: sheet '{sheet_name}' not found. "
                    f"Available: {', '.join(wb.sheetnames)}"
                )
            ws = wb[sheet_name]
        else:
            ws = wb.active

        rows_iter = ws.iter_rows(values_only=True)

        # Skip rows before header
        for _ in range(skip_rows):
            try:
                next(rows_iter)
            except StopIteration:
                wb.close()
                raise ValueError("Excel Source: not enough rows after skip_rows offset")

        # Read header row
        header_rows_to_skip = max(0, header_row - 1 - skip_rows)
        for _ in range(header_rows_to_skip):
            try:
                next(rows_iter)
            except StopIteration:
                wb.close()
                raise ValueError("Excel Source: header row beyond available data")

        try:
            header = next(rows_iter)
        except StopIteration:
            wb.close()
            raise ValueError("Excel Source: no data found in the worksheet")

        columns = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(header)]

        # Read data rows
        data_rows = []
        for row in rows_iter:
            if all(cell is None for cell in row):
                continue  # Skip fully empty rows
            data_rows.append(row)

        wb.close()

        if not data_rows:
            col_defs = ", ".join(f"NULL AS \"{c}\"" for c in columns)
            return conn.sql(f"SELECT {col_defs} WHERE false")

        # Build VALUES clause
        def fmt(v):
            if v is None:
                return "NULL"
            if isinstance(v, bool):
                return "true" if v else "false"
            if isinstance(v, (int, float)):
                return str(v)
            return "'" + str(v).replace("'", "''") + "'"

        value_parts = []
        for row in data_rows:
            vals = ", ".join(fmt(v) for v in row)
            value_parts.append(f"({vals})")

        values_sql = ", ".join(value_parts)
        _qcols = ", ".join(f'"{c}"' for c in columns)
        # Unique per-call temp-table name so two Excel-source nodes in one
        # pipeline don't clobber each other's staged rows (the returned relation
        # is lazy over this table).
        import uuid as _uuid
        tmp = f"__excel_source_{_uuid.uuid4().hex[:8]}"
        conn.execute(
            f"CREATE OR REPLACE TEMP TABLE {tmp} "
            f"AS SELECT * FROM (VALUES {values_sql}) AS __vals ({_qcols})"
        )
        return conn.sql(f"SELECT * FROM {tmp}")

    @staticmethod
    def default_params() -> dict[str, Any]:
        return {"file_path": "", "sheet_name": "", "header_row": 1, "skip_rows": 0}

    @staticmethod
    def param_schema() -> list[dict]:
        return [
            {"name": "file_path", "type": "file", "label": "File Path", "required": True,
             "placeholder": "data/report.xlsx"},
            {"name": "sheet_name", "type": "text", "label": "Sheet Name",
             "placeholder": "Sheet1",
             "description": "Leave empty for the active (first) sheet"},
            {"name": "header_row", "type": "number", "label": "Header Row", "default": 1,
             "description": "Row number containing column headers (1-based)"},
            {"name": "skip_rows", "type": "number", "label": "Skip Rows", "default": 0,
             "description": "Number of rows to skip from the top before the header"},
        ]


# ── XML Source ──

@register(StepType.XML_SOURCE)
class XmlSourceNode(BaseNode):
    """Read data from XML files by parsing elements into rows."""
    display_name = "XML Source"
    category = "source"
    description = "Read data from an XML file"

    def execute(self, ctx: ExecutionContext) -> duckdb.DuckDBPyRelation:
        file_path = _resolve_path(self.params.get("file_path", ""), ctx.data_dir)

        if not os.path.isfile(file_path):
            raise ValueError(f"XML Source: file not found: {file_path}")

        row_tag = self.params.get("row_tag", "")
        encoding = self.params.get("encoding", "utf-8")

        import xml.etree.ElementTree as ET

        tree = ET.parse(file_path)
        root = tree.getroot()

        # Strip namespace prefixes for easier element access
        def strip_ns(tag: str) -> str:
            return tag.split("}")[-1] if "}" in tag else tag

        # Find row elements
        if row_tag:
            # Search recursively for elements matching row_tag
            elements = root.iter()
            row_elements = [el for el in elements if strip_ns(el.tag) == row_tag]
        else:
            # Auto-detect: use direct children of root
            row_elements = list(root)

        if not row_elements:
            raise ValueError(
                f"XML Source: no elements found"
                + (f" with tag '{row_tag}'" if row_tag else "")
                + ". Specify row_tag parameter."
            )

        # Convert each element to a dict
        rows: list[dict] = []
        for elem in row_elements:
            row: dict[str, Any] = {}
            # Include element attributes
            for attr_name, attr_val in elem.attrib.items():
                row[strip_ns(attr_name)] = attr_val
            # Include child element text
            for child in elem:
                tag = strip_ns(child.tag)
                text = (child.text or "").strip()
                # If child has its own children, serialize as string
                if len(child) > 0:
                    text = ET.tostring(child, encoding="unicode", method="text").strip()
                row[tag] = text if text else None
            # If element has direct text and no children, store as "value"
            if not row and elem.text and elem.text.strip():
                row["value"] = elem.text.strip()
            if row:
                rows.append(row)

        if not rows:
            raise ValueError("XML Source: parsed 0 rows from the XML file")

        return _rows_to_relation(ctx.conn, rows, "__xml_source")

    @staticmethod
    def default_params() -> dict[str, Any]:
        return {"file_path": "", "row_tag": "", "encoding": "utf-8"}

    @staticmethod
    def param_schema() -> list[dict]:
        return [
            {"name": "file_path", "type": "file", "label": "File Path", "required": True,
             "placeholder": "data/input.xml"},
            {"name": "row_tag", "type": "text", "label": "Row Element Tag",
             "placeholder": "record",
             "description": "XML tag name for each row. Leave empty to use root's children."},
            {"name": "encoding", "type": "select", "label": "Encoding",
             "options": ["utf-8", "utf-16", "latin-1", "ascii"], "default": "utf-8"},
        ]


# ── S3/MinIO Source ──

@register(StepType.S3_SOURCE)
class S3SourceNode(BaseNode):
    """Read files from S3 or MinIO object storage."""
    display_name = "S3 / MinIO Source"
    category = "source"
    description = "Read data from S3 or MinIO object storage"

    def execute(self, ctx: ExecutionContext) -> duckdb.DuckDBPyRelation:
        bucket = self.params.get("bucket", "")
        key = self.params.get("key", "")
        endpoint = self.params.get("endpoint", "")
        access_key = self.params.get("access_key", "")
        secret_key = self.params.get("secret_key", "")
        region = self.params.get("region", "us-east-1")
        file_format = self.params.get("format", "auto")
        connection_id = self.params.get("connection_id", "")

        # Try loading from saved connection
        if connection_id:
            from fpulse.nodes.db_source import _get_connection_config
            result = _get_connection_config(connection_id)
            if result:
                config, _ = result
                endpoint = config.get("endpoint", endpoint)
                access_key = config.get("access_key", access_key)
                secret_key = config.get("secret_key", secret_key)
                region = config.get("region", region)
                bucket = config.get("bucket", bucket)

        if not bucket or not key:
            raise ValueError("S3 Source: bucket and key are required")

        # Try boto3 first
        try:
            import boto3
            return self._read_with_boto3(
                ctx, bucket, key, endpoint, access_key, secret_key, region, file_format
            )
        except ImportError:
            pass

        # Fallback: httpx/urllib with presigned-style URL
        return self._read_with_http(ctx, bucket, key, endpoint, file_format)

    def _read_with_boto3(self, ctx: ExecutionContext, bucket: str, key: str,
                          endpoint: str, access_key: str, secret_key: str,
                          region: str, file_format: str) -> duckdb.DuckDBPyRelation:
        """Read S3 object using boto3."""
        import boto3

        client_kwargs: dict[str, Any] = {"region_name": region}
        if endpoint:
            client_kwargs["endpoint_url"] = endpoint
        if access_key and secret_key:
            client_kwargs["aws_access_key_id"] = access_key
            client_kwargs["aws_secret_access_key"] = secret_key

        s3 = boto3.client("s3", **client_kwargs)

        # Download to temp file
        suffix = os.path.splitext(key)[-1] or ".tmp"
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
            s3.download_fileobj(bucket, key, tmp)
            tmp_path = tmp.name

        try:
            return self._read_file(ctx, tmp_path, key, file_format)
        finally:
            try:
                os.unlink(tmp_path)
            except OSError:
                pass

    def _read_with_http(self, ctx: ExecutionContext, bucket: str, key: str,
                         endpoint: str, file_format: str) -> duckdb.DuckDBPyRelation:
        """Read S3/MinIO object via HTTP (unsigned or public access)."""
        if endpoint:
            url = f"{endpoint.rstrip('/')}/{bucket}/{key}"
        else:
            url = f"https://{bucket}.s3.amazonaws.com/{key}"

        suffix = os.path.splitext(key)[-1] or ".tmp"
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
            try:
                req = urllib.request.Request(url, headers={"User-Agent": "F-Pulse/0.6.0"})
                with urllib.request.urlopen(req, timeout=60) as resp:
                    tmp.write(resp.read())
            except urllib.error.URLError as exc:
                raise ValueError(
                    f"S3 Source: failed to download s3://{bucket}/{key}: {exc.reason}. "
                    f"Install boto3 for authenticated access: pip install boto3"
                ) from exc
            tmp_path = tmp.name

        try:
            return self._read_file(ctx, tmp_path, key, file_format)
        finally:
            try:
                os.unlink(tmp_path)
            except OSError:
                pass

    @staticmethod
    def _read_file(ctx: ExecutionContext, local_path: str, key: str,
                    file_format: str) -> duckdb.DuckDBPyRelation:
        """Read a local file based on format (auto-detect from extension)."""
        if file_format == "auto":
            ext = os.path.splitext(key)[-1].lower()
            fmt_map = {
                ".csv": "csv", ".tsv": "csv", ".json": "json",
                ".parquet": "parquet", ".pq": "parquet",
            }
            file_format = fmt_map.get(ext, "csv")

        if file_format == "parquet":
            return ctx.conn.read_parquet(local_path)
        elif file_format == "json":
            return ctx.conn.read_json(local_path)
        else:
            delimiter = "\t" if key.endswith(".tsv") else ","
            return ctx.conn.read_csv(local_path, delimiter=delimiter, header=True)

    @staticmethod
    def default_params() -> dict[str, Any]:
        return {
            "connection_id": "", "bucket": "", "key": "", "endpoint": "",
            "access_key": "", "secret_key": "", "region": "us-east-1", "format": "auto",
        }

    @staticmethod
    def param_schema() -> list[dict]:
        return [
            {"name": "connection_id", "type": "connection_picker", "label": "Connection",
             "description": "Select a saved S3/MinIO connection, or fill fields below."},
            {"name": "bucket", "type": "text", "label": "Bucket", "required": True,
             "placeholder": "my-data-bucket"},
            {"name": "key", "type": "text", "label": "Object Key", "required": True,
             "placeholder": "raw/2024/sales.parquet"},
            {"name": "endpoint", "type": "text", "label": "Endpoint URL",
             "placeholder": "http://localhost:9000",
             "description": "Required for MinIO. Leave empty for AWS S3."},
            {"name": "access_key", "type": "text", "label": "Access Key",
             "description": "Overrides connection credential if set."},
            {"name": "secret_key", "type": "password", "label": "Secret Key"},
            {"name": "region", "type": "text", "label": "Region", "default": "us-east-1"},
            {"name": "format", "type": "select", "label": "File Format",
             "options": ["auto", "csv", "json", "parquet"], "default": "auto",
             "description": "Auto-detects from file extension."},
            # X3 (2026-05-30) — sync_mode declarative marker. S3 listings
            # support modified_after server-side via the LIST API's
            # `--query` filter (boto3) — the operator embeds {cursor}
            # in the `key` glob until the engine wires it natively.
            *__import__("fpulse.nodes._sync_mode_decl",
                        fromlist=["sync_mode_marker_entries"]).sync_mode_marker_entries(
                "Embed {cursor} in the object key glob (e.g. "
                "`raw/{cursor}/*.parquet`) — the engine substitutes "
                "the last persisted cursor before listing. The auto-save "
                "stamps the new high-water-mark as MAX of the listed "
                "object keys.",
            ),
        ]


# ── Kafka Source ──

@register(StepType.KAFKA_SOURCE)
class KafkaSourceNode(BaseNode):
    """Read messages from a Kafka topic."""
    display_name = "Kafka Source"
    category = "source"
    description = "Read messages from a Kafka topic"

    def execute(self, ctx: ExecutionContext) -> duckdb.DuckDBPyRelation:
        topic = self.params.get("topic", "")
        bootstrap_servers = self.params.get("bootstrap_servers", "localhost:9092")
        group_id = self.params.get("group_id", "fpulse-consumer")
        max_messages = self.params.get("max_messages", 100)
        timeout_seconds = self.params.get("timeout_seconds", 10)
        connection_id = self.params.get("connection_id", "")

        if not topic:
            raise ValueError("Kafka Source: topic is required")

        # Try loading connection config
        if connection_id:
            from fpulse.nodes.db_source import _get_connection_config
            result = _get_connection_config(connection_id)
            if result:
                config, _ = result
                bootstrap_servers = config.get("bootstrap_servers", bootstrap_servers)
                group_id = config.get("group_id", group_id)

        # Try confluent_kafka first, then kafka-python
        try:
            return self._read_confluent(
                ctx, topic, bootstrap_servers, group_id, max_messages, timeout_seconds
            )
        except ImportError:
            pass

        try:
            return self._read_kafka_python(
                ctx, topic, bootstrap_servers, group_id, max_messages, timeout_seconds
            )
        except ImportError:
            raise ValueError(
                "Kafka Source: no Kafka client library installed. "
                "Install one: pip install confluent-kafka  OR  pip install kafka-python"
            )

    def _read_confluent(self, ctx: ExecutionContext, topic: str,
                         bootstrap_servers: str, group_id: str,
                         max_messages: int,
                         timeout_seconds: int) -> duckdb.DuckDBPyRelation:
        """Read using confluent_kafka."""
        from confluent_kafka import Consumer  # type: ignore

        conf = {
            "bootstrap.servers": bootstrap_servers,
            "group.id": group_id,
            "auto.offset.reset": "earliest",
            "enable.auto.commit": False,
        }
        consumer = Consumer(conf)
        consumer.subscribe([topic])

        rows: list[dict] = []
        try:
            deadline = __import__("time").time() + timeout_seconds
            while len(rows) < max_messages and __import__("time").time() < deadline:
                msg = consumer.poll(timeout=1.0)
                if msg is None:
                    continue
                if msg.error():
                    continue
                value = msg.value()
                if value is None:
                    continue
                decoded = value.decode("utf-8", errors="replace")
                try:
                    row = json.loads(decoded)
                    if isinstance(row, dict):
                        row["_kafka_offset"] = msg.offset()
                        row["_kafka_partition"] = msg.partition()
                        row["_kafka_topic"] = msg.topic()
                        rows.append(row)
                    else:
                        rows.append({
                            "value": decoded,
                            "_kafka_offset": msg.offset(),
                            "_kafka_partition": msg.partition(),
                            "_kafka_topic": msg.topic(),
                        })
                except json.JSONDecodeError:
                    rows.append({
                        "value": decoded,
                        "_kafka_offset": msg.offset(),
                        "_kafka_partition": msg.partition(),
                        "_kafka_topic": msg.topic(),
                    })
        finally:
            consumer.close()

        if not rows:
            return ctx.conn.sql(
                "SELECT NULL::VARCHAR AS value, NULL::BIGINT AS _kafka_offset, "
                "NULL::INT AS _kafka_partition, NULL::VARCHAR AS _kafka_topic WHERE false"
            )

        return _rows_to_relation(ctx.conn, rows, "__kafka_source")

    def _read_kafka_python(self, ctx: ExecutionContext, topic: str,
                            bootstrap_servers: str, group_id: str,
                            max_messages: int,
                            timeout_seconds: int) -> duckdb.DuckDBPyRelation:
        """Read using kafka-python."""
        from kafka import KafkaConsumer  # type: ignore

        consumer = KafkaConsumer(
            topic,
            bootstrap_servers=bootstrap_servers.split(","),
            group_id=group_id,
            auto_offset_reset="earliest",
            enable_auto_commit=False,
            consumer_timeout_ms=timeout_seconds * 1000,
            value_deserializer=lambda v: v.decode("utf-8", errors="replace"),
        )

        rows: list[dict] = []
        try:
            for msg in consumer:
                try:
                    row = json.loads(msg.value)
                    if isinstance(row, dict):
                        row["_kafka_offset"] = msg.offset
                        row["_kafka_partition"] = msg.partition
                        row["_kafka_topic"] = msg.topic
                        rows.append(row)
                    else:
                        rows.append({
                            "value": msg.value,
                            "_kafka_offset": msg.offset,
                            "_kafka_partition": msg.partition,
                            "_kafka_topic": msg.topic,
                        })
                except json.JSONDecodeError:
                    rows.append({
                        "value": msg.value,
                        "_kafka_offset": msg.offset,
                        "_kafka_partition": msg.partition,
                        "_kafka_topic": msg.topic,
                    })
                if len(rows) >= max_messages:
                    break
        finally:
            consumer.close()

        if not rows:
            return ctx.conn.sql(
                "SELECT NULL::VARCHAR AS value, NULL::BIGINT AS _kafka_offset, "
                "NULL::INT AS _kafka_partition, NULL::VARCHAR AS _kafka_topic WHERE false"
            )

        return _rows_to_relation(ctx.conn, rows, "__kafka_source")

    @staticmethod
    def default_params() -> dict[str, Any]:
        return {
            "connection_id": "", "topic": "", "bootstrap_servers": "localhost:9092",
            "group_id": "fpulse-consumer", "max_messages": 100, "timeout_seconds": 10,
        }

    @staticmethod
    def param_schema() -> list[dict]:
        return [
            {"name": "connection_id", "type": "connection_picker", "label": "Connection",
             "description": "Select a saved Kafka connection, or fill fields below."},
            {"name": "topic", "type": "text", "label": "Topic", "required": True,
             "placeholder": "events.raw"},
            {"name": "bootstrap_servers", "type": "text", "label": "Bootstrap Servers",
             "default": "localhost:9092",
             "description": "Comma-separated broker addresses."},
            {"name": "group_id", "type": "text", "label": "Consumer Group",
             "default": "fpulse-consumer"},
            {"name": "max_messages", "type": "number", "label": "Max Messages",
             "default": 100, "description": "Maximum number of messages to consume."},
            {"name": "timeout_seconds", "type": "number", "label": "Timeout (seconds)",
             "default": 10},
        ]


# ── FTP Source ──

@register(StepType.FTP_SOURCE)
class FtpSourceNode(BaseNode):
    """Read files from an FTP, FTPS, or SFTP server.

    FTP/FTPS use ``ftplib`` (stdlib); SFTP (SSH File Transfer Protocol — a
    DIFFERENT protocol) uses ``paramiko`` (lazy-imported, with a clear install
    hint if missing). The file is downloaded to a temp path then read by the
    shared file reader, so all three protocols support csv/json/parquet/etc.
    """
    display_name = "FTP / SFTP Source"
    category = "source"
    description = "Read data from an FTP, FTPS, or SFTP server"

    @staticmethod
    def _resolve_protocol(params: dict[str, Any]) -> str:
        """ftp | ftps | sftp. Explicit `protocol` wins; else inferred from the
        generic connector_type (sftp) or the legacy use_tls flag (ftps)."""
        p = (params.get("protocol") or "").strip().lower()
        if p in ("ftp", "ftps", "sftp"):
            return p
        if (params.get("connector_type") or "").strip().lower() == "sftp":
            return "sftp"
        return "ftps" if params.get("use_tls") else "ftp"

    def execute(self, ctx: ExecutionContext) -> duckdb.DuckDBPyRelation:
        host = self.params.get("host", "")
        port = self.params.get("port")
        username = self.params.get("username", "anonymous")
        password = self.params.get("password", "")
        remote_path = self.params.get("remote_path", "")
        file_format = self.params.get("format", "auto")
        private_key = self.params.get("private_key", "")
        connection_id = self.params.get("connection_id", "")
        protocol = self._resolve_protocol(self.params)

        host_key_config: dict[str, Any] = dict(self.params)
        if connection_id:
            from fpulse.nodes.db_source import _get_connection_config
            result = _get_connection_config(connection_id)
            if result:
                config, _ = result
                host_key_config.update(config)
                host = config.get("host", host)
                port = config.get("port", port)
                username = config.get("username", username)
                password = config.get("password", password)
                private_key = config.get("private_key", private_key)
                if config.get("protocol"):
                    protocol = str(config["protocol"]).strip().lower()

        if not host:
            raise ValueError("FTP/SFTP Source: host is required")
        if not remote_path:
            raise ValueError("FTP/SFTP Source: remote_path is required")

        # Effective port: explicit value, else the protocol default. If the
        # user left the FTP default 21 but chose SFTP, use 22.
        try:
            eff_port = int(port) if port not in (None, "", 0) else 0
        except (TypeError, ValueError):
            eff_port = 0
        if eff_port == 0:
            eff_port = 22 if protocol == "sftp" else 21
        elif protocol == "sftp" and eff_port == 21:
            eff_port = 22

        suffix = os.path.splitext(remote_path)[-1] or ".tmp"

        if protocol == "sftp":
            try:
                import paramiko
            except ImportError as exc:
                raise ValueError(
                    "SFTP requires the 'paramiko' package — install it with "
                    "`pip install paramiko`, then retry. (FTP/FTPS need no extra "
                    "package.)"
                ) from exc
            from fpulse.nodes._sftp import build_ssh_client
            ssh = build_ssh_client(host_key_config)
            ckwargs: dict[str, Any] = {
                "hostname": host, "port": eff_port, "username": username, "timeout": 30,
            }
            pkey = None
            if private_key:
                from io import StringIO
                for kcls in (paramiko.RSAKey, paramiko.Ed25519Key, paramiko.ECDSAKey):
                    try:
                        pkey = kcls.from_private_key(StringIO(str(private_key)))
                        break
                    except Exception:  # noqa: BLE001 — try the next key type
                        continue
            if pkey is not None:
                ckwargs["pkey"] = pkey
            elif password:
                ckwargs["password"] = password
            ssh.connect(**ckwargs)
            try:
                with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
                    tmp_path = tmp.name
                ssh.open_sftp().get(remote_path, tmp_path)
            finally:
                ssh.close()
        else:
            import ftplib
            ftp_class = ftplib.FTP_TLS if protocol == "ftps" else ftplib.FTP
            ftp = ftp_class()
            ftp.connect(host, eff_port, timeout=30)
            ftp.login(username, password)
            if protocol == "ftps" and isinstance(ftp, ftplib.FTP_TLS):
                ftp.prot_p()
            with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
                ftp.retrbinary(f"RETR {remote_path}", tmp.write)
                tmp_path = tmp.name
            ftp.quit()

        try:
            return S3SourceNode._read_file(ctx, tmp_path, remote_path, file_format)
        finally:
            try:
                os.unlink(tmp_path)
            except OSError:
                pass

    @staticmethod
    def default_params() -> dict[str, Any]:
        return {
            "connection_id": "", "host": "", "protocol": "ftp", "port": 21,
            "username": "anonymous", "password": "", "private_key": "",
            "remote_path": "", "format": "auto",
        }

    @staticmethod
    def param_schema() -> list[dict]:
        return [
            {"name": "connection_id", "type": "connection_picker", "label": "Connection",
             "description": "Select a saved FTP/SFTP connection, or fill fields below."},
            {"name": "host", "type": "text", "label": "Host", "required": True,
             "placeholder": "sftp.example.com"},
            {"name": "protocol", "type": "select", "label": "Protocol",
             "options": ["ftp", "ftps", "sftp"], "default": "ftp",
             "description": "sftp = SSH File Transfer (needs paramiko); ftps = FTP over TLS."},
            {"name": "port", "type": "number", "label": "Port", "default": 21,
             "description": "FTP/FTPS = 21, SFTP = 22 (auto-corrected if left at 21)."},
            {"name": "username", "type": "text", "label": "Username", "default": "anonymous"},
            {"name": "password", "type": "password", "label": "Password"},
            {"name": "private_key", "type": "password", "label": "Private Key (SFTP, optional)",
             "description": "PEM private key for SFTP key-based auth — alternative to password."},
            {"name": "remote_path", "type": "text", "label": "Remote File Path", "required": True,
             "placeholder": "/data/exports/report.csv"},
            {"name": "format", "type": "select", "label": "File Format",
             "options": ["auto", "csv", "json", "parquet"], "default": "auto"},
        ]


# ── Google Sheets Source ──

@register(StepType.GSHEET_SOURCE)
class GSheetSourceNode(BaseNode):
    """Read data from a public Google Sheet via its CSV export URL."""
    display_name = "Google Sheets Source"
    category = "source"
    description = "Read data from a public Google Sheet"

    def execute(self, ctx: ExecutionContext) -> duckdb.DuckDBPyRelation:
        sheet_url = self.params.get("sheet_url", "")
        sheet_id = self.params.get("sheet_id", "")
        gid = self.params.get("gid", "0")

        if not sheet_url and not sheet_id:
            raise ValueError(
                "Google Sheets Source: provide either a sheet_url or sheet_id"
            )

        # Extract sheet_id from URL if full URL provided
        if sheet_url and not sheet_id:
            # URL format: https://docs.google.com/spreadsheets/d/SHEET_ID/...
            import re
            match = re.search(r"/spreadsheets/d/([a-zA-Z0-9_-]+)", sheet_url)
            if not match:
                raise ValueError(
                    "Google Sheets Source: could not extract sheet ID from URL. "
                    "Expected format: https://docs.google.com/spreadsheets/d/SHEET_ID/..."
                )
            sheet_id = match.group(1)

            # Try to extract gid from URL
            gid_match = re.search(r"[?&]gid=(\d+)", sheet_url)
            if gid_match:
                gid = gid_match.group(1)

        # Build CSV export URL
        csv_url = (
            f"https://docs.google.com/spreadsheets/d/{sheet_id}"
            f"/export?format=csv&gid={gid}"
        )

        # Download CSV
        with tempfile.NamedTemporaryFile(delete=False, suffix=".csv", mode="wb") as tmp:
            try:
                req = urllib.request.Request(
                    csv_url,
                    headers={"User-Agent": "F-Pulse/0.6.0"},
                )
                with urllib.request.urlopen(req, timeout=30) as resp:
                    content_type = resp.headers.get("Content-Type", "")
                    if "text/html" in content_type:
                        raise ValueError(
                            "Google Sheets Source: the sheet is not publicly accessible. "
                            "Set sharing to 'Anyone with the link' in Google Sheets."
                        )
                    tmp.write(resp.read())
            except urllib.error.HTTPError as exc:
                raise ValueError(
                    f"Google Sheets Source: HTTP {exc.code} — "
                    f"the sheet may not be publicly shared."
                ) from exc
            except urllib.error.URLError as exc:
                raise ValueError(
                    f"Google Sheets Source: network error: {exc.reason}"
                ) from exc
            tmp_path = tmp.name

        try:
            return ctx.conn.read_csv(tmp_path, header=True)
        finally:
            try:
                os.unlink(tmp_path)
            except OSError:
                pass

    @staticmethod
    def default_params() -> dict[str, Any]:
        return {"sheet_url": "", "sheet_id": "", "gid": "0"}

    @staticmethod
    def param_schema() -> list[dict]:
        return [
            {"name": "sheet_url", "type": "text", "label": "Google Sheet URL",
             "placeholder": "https://docs.google.com/spreadsheets/d/.../edit",
             "description": "Full URL to the Google Sheet (must be publicly shared)."},
            {"name": "sheet_id", "type": "text", "label": "Sheet ID",
             "placeholder": "1BxiMVs0XRA5nFMdKvBdBZjgmUUqptlbs74OgVE2upms",
             "description": "Alternative: just the sheet ID from the URL."},
            {"name": "gid", "type": "text", "label": "Tab GID", "default": "0",
             "description": "Sheet tab ID (0 = first tab). Found in the URL as gid=..."},
            # X3 (2026-05-30) — declarative sync_mode marker. Google
            # Sheets is append-only in most ops use; the cursor is
            # typically a row-id or last-edited timestamp column the
            # operator filters on post-fetch via a downstream Filter
            # node. Auto-substitution is roadmap.
            *__import__("fpulse.nodes._sync_mode_decl",
                        fromlist=["sync_mode_marker_entries"]).sync_mode_marker_entries(
                "Google Sheets has no server-side cursor filter. Read "
                "the full sheet here, then add a downstream Filter node "
                "with `row_id > {cursor}` (the engine substitutes the "
                "persisted cursor into the Filter expression). Auto-save "
                "stamps MAX(row_id) from the result.",
            ),
        ]


# ── Delta Lake Source ──

@register(StepType.DELTA_SOURCE)
class DeltaSourceNode(BaseNode):
    """Read Delta Lake tables via their Parquet data files."""
    display_name = "Delta Lake Source"
    category = "source"
    description = "Read a Delta Lake table"

    def execute(self, ctx: ExecutionContext) -> duckdb.DuckDBPyRelation:
        table_path = self.params.get("table_path", "")
        if not table_path:
            raise ValueError("Delta Source: table_path is required")

        if not os.path.isabs(table_path):
            table_path = os.path.join(ctx.data_dir, table_path)

        version = self.params.get("version", None)
        columns = self.params.get("columns", [])

        # Try deltalake library first (proper Delta protocol support)
        try:
            import deltalake  # type: ignore
            return self._read_with_deltalake(ctx, table_path, version, columns)
        except ImportError:
            pass

        # Fallback: read Parquet files from the Delta directory
        return self._read_parquet_fallback(ctx, table_path, columns)

    @staticmethod
    def _read_with_deltalake(ctx: ExecutionContext, table_path: str,
                              version: int | None,
                              columns: list[str]) -> duckdb.DuckDBPyRelation:
        """Read using the deltalake Python library."""
        import deltalake  # type: ignore

        dt_kwargs: dict[str, Any] = {}
        if version is not None:
            dt_kwargs["version"] = int(version)

        dt = deltalake.DeltaTable(table_path, **dt_kwargs)
        arrow_table = dt.to_pyarrow_table(columns=columns if columns else None)
        return ctx.conn.from_arrow(arrow_table)

    @staticmethod
    def _read_parquet_fallback(ctx: ExecutionContext, table_path: str,
                                columns: list[str]) -> duckdb.DuckDBPyRelation:
        """Fallback: read Parquet files directly from Delta directory."""
        import glob as glob_mod

        # Delta tables store data in Parquet files at the table root
        parquet_pattern = os.path.join(table_path, "*.parquet")
        parquet_files = glob_mod.glob(parquet_pattern)

        # Also check for partitioned data in subdirectories
        if not parquet_files:
            parquet_pattern = os.path.join(table_path, "**", "*.parquet")
            parquet_files = glob_mod.glob(parquet_pattern, recursive=True)

        if not parquet_files:
            raise ValueError(
                f"Delta Source: no Parquet files found in {table_path}. "
                f"Install deltalake for proper Delta protocol support: "
                f"pip install deltalake"
            )

        # Use DuckDB to read all matching Parquet files
        pattern_for_duckdb = os.path.join(table_path, "**", "*.parquet")
        if columns:
            col_list = ", ".join(f'"{c}"' for c in columns)
            return ctx.conn.sql(
                f"SELECT {col_list} FROM read_parquet('{pattern_for_duckdb}', "
                f"hive_partitioning=true)"
            )

        return ctx.conn.sql(
            f"SELECT * FROM read_parquet('{pattern_for_duckdb}', "
            f"hive_partitioning=true)"
        )

    @staticmethod
    def default_params() -> dict[str, Any]:
        return {"table_path": "", "version": None, "columns": []}

    @staticmethod
    def param_schema() -> list[dict]:
        return [
            {"name": "table_path", "type": "text", "label": "Table Path", "required": True,
             "placeholder": "lakehouse/bronze/sales",
             "description": "Path to the Delta Lake table directory."},
            {"name": "version", "type": "number", "label": "Version (Time Travel)",
             "description": "Read a specific version of the table. Leave empty for latest."},
            {"name": "columns", "type": "column_list", "label": "Select Columns",
             "description": "Leave empty to read all columns."},
        ]
